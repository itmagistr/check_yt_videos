# -*- coding: utf-8 -*-
import os
import sys
#import json
import msvcrt
import statistics
import traceback
import argparse
from pathlib import Path
import time
import datetime
import openpyxl
import pyperclip
import googleapiclient.errors      #список ошибок доступа к YouTube Data API
import googleapiclient.discovery   #клиент доступа к YouTube Data API
from check_yt_models import *
from pony import orm
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import * #
import logging
logging.basicConfig()
logging.getLogger().setLevel(logging.INFO)
logging.getLogger('googleapicliet.discovery_cache').setLevel(logging.CRITICAL)
logging.getLogger().addHandler(logging.FileHandler("yt_optima.log"))

P100ms = 0.20    # 0.05-0.20 - базовая пауза. Остальные паузы указаны кратными базовой
PMIN = 10*P100ms # 10* - минимальная пауза для прорисовки пользовательского интерфейса
MinScore =  0.8   # минимальное  значение тега для включения в рабочее облако
WrkScore =  6.8   # начало рабочего диапазона тегов, лучше привести к 7.1
MaxScore = 32.0   # максимальное значение тега для включения в рабочее облако

FLTYPE={'Обзор': {'litera': 'ov',},
		'Просмотры': {'litera': 'v',},
		'Взаимодействие': {'litera': 'iact',},
		'Аудитория': {'litera': 'aud',},
		}

def main(opts):
	logging.info(f'Начало работы главного модуля')
	if len(opts.chID) > 1:
		logging.info(f'Начало формирования списка видео канала {opts.chID}')
		getVideoList(opts) # сформировать файл ссылок на видео канала

		return #выйти из выполнения

	urls=loadUrls(opts.infile)
	urlslen = len(urls.keys())
	logging.info('Подготовлено {} видео для анализа'.format(urlslen))
	addTags = loadTags(opts.addtags)
	logging.info('Подготовлено {} тегов для анализа'.format(len(addTags)))
	driver = connect2Browser(opts)

	if opts.words=='1':
		# собрать рейтинг seo для слов из файла для видео в первой строке
		rateWords(driver, opts)
	else:
		excludeTags = []
		if opts.update=='2':
			excludeTags = getExtags(1, 2, 6.81) # min< 1, avg< 2, max< 7
		crow=0
		for i, u in urls.items():
			crow+=1
			exitOnKey() # key q
			logging.info('-------- Обработка {:03d} из {:03d} ({:06.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen, u))
			checkPauseKey() # key p
			cdata = {}
			
			if opts.update == '2':
				# для видео добавить теги, которые есть в БД, но не проверялись под данным видео
				for t in getDBtags(i, u, excludeTags):
					cdata[t]={}
				if len(cdata.keys()) == 0:
					logging.info('Для видео все доступные теги оценены')
					continue
				else:
					logging.info('Для видео отобраны теги для оценки {}'.format(len(cdata.keys())))
			if opts.update == '3':
				# обновление тегов если значение рейтингов выше
				odt = datetime.datetime.strptime(opts.dt, '%Y-%m-%d %H:%M')
				d = CheckSavedData(i, odt)
				if d is not None:
					logging.info('treal:{} -> {}, tshow: {} -> {}, at {}'.format(d['rate'][0]['treal'], d['rate'][1]['treal'], d['rate'][0]['tshow'], d['rate'][1]['tshow'], d['dt']))
				else:
					tagsUpdate_V3(driver, i, u, opts)
					#tagsUpdate(driver, i, u, opts) # первая попытка

			elif opts.update in '0124': #режим 0, 1, 2, 4
				driver.get(u)
				time.sleep(P100ms*20) # пауза 2 при заходе на очередное видео
				try:
					testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
				except TimeoutException:
					logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
					continue
				
				inpts = driver.find_elements_by_xpath("//ytcp-mention-input")
				vid_title = inpts[0].text
				
				if opts.update == '4':
					# для видео добавить теги, которые входят в название и еще не проверялись под данным видео
					inpts = driver.find_elements_by_xpath("//ytcp-mention-input")
					vid_title = inpts[0].text
					cnt = 0
					for t in getTitleTags(i, u, vid_title, addTags):
						cdata[t]={}
						cnt+=1
					logging.info(f'тегов {cnt} для оценки с учетом названия "{vid_title}"')
				elif opts.update in '01': #режим 0 или 1
					#собрать теги
					for elm in driver.find_elements_by_xpath("//*[@id='child-input']//ytcp-chip[@role='button']"):
						text = elm.get_attribute('vidiq-keyword')
						if opts.update == '0':
							# все теги видео добавить к проверке
							cdata[text]={}
						elif opts.update == '1':
							if getDBTagSeo(i, u, text) < 0.01:
								# новые и с оценкой ноль добавить для проверки
								cdata[text]={}

				
				# добавить два слова для очистки тегов, чтобы гарантированно отображалась кнопка удалить все теги
				elq = driver.find_element_by_id('text-input')
				elq.click()
				
				# #clear all tags
				doit_clear=2
				while doit_clear>0:
					try:
						di = driver.find_element_by_id('clear-button')
						di.click()
						doit_clear=0
					except:
						logging.info("find_element_by_id('clear-button') Unexpected error: {}".format(sys.exc_info()[0]))
						elq.send_keys('wr1,wr2', Keys.ENTER)
						time.sleep(PMIN)
						doit_clear-=1

				# по тегам сохранить рейтинг
				tagscnt=1+len(cdata.keys())
				for t in cdata.keys():
					tagscnt-=1
					#add new tags
					elq.click()
					if opts.clipboard == '1':
						pyperclip.copy(t)
						time.sleep(PMIN)
						#elq.send_keys(Keys.CONTROL, 'v')
						elq.send_keys(Keys.SHIFT, Keys.INSERT)
					else:
						elq.send_keys(t, Keys.ENTER)
					time.sleep(P100ms*30) # пауза 3 для обновления значения VidIQ

					#get SEO score
					elm=driver.find_element_by_class_name('stat-value-seo-score')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					#logging.info('seo {}'.format(elm1.text))
					cdata[t]['seo'] = float(elm1.text)
					
					if cdata[t]['seo'] < 0.01:
						# если не успело значение обновиться, то подождем еще и вычитаем повторно
						time.sleep(P100ms*20) # 2 пауза второго шанса обновления значения VidIQ
						#get SEO score
						elm=driver.find_element_by_class_name('stat-value-seo-score')
						elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
						#logging.info('seo2 {}'.format(elm1.text))
						cdata[t]['seo'] = float(elm1.text)
					getScores(driver, cdata[t])
					
					# удалить тег
					try:
						btn = driver.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
						if len(btn) > 1:
							di = driver.find_element_by_id('clear-button')
							di.click()
						else:
							btn[0].click()
					except:
						logging.info("ERROR: {}".format(sys.exc_info()[0]))
						
					extstr = ''
					if cdata[t]['ranked'] > 0:
						# если тег ранжированный, то к оценке добавим 0.01
						cdata[t]['seo'] += 0.01
						extstr = '< РАНЖИРОВАННЫЙ >'
					if t.upper() in vid_title.upper():
						# если тег входит в строку названия, то к оценке добавим 0.02
						cdata[t]['seo'] += 0.02
					saveindb({'vid': i, 'url': u, 'tags': {t: cdata[t]}, })
					logging.info('{:03d}/{:03d}, {:03d} {}->{} {}'.format(crow, urlslen, tagscnt, t, cdata[t]['seo'], extstr))
					checkPauseKey()

				#logging.info('cancel changes')
				btn=driver.find_element_by_id('discard')
				btn.click()
				time.sleep(PMIN)

				# if opts.update!='2':
				# 	saveindb({'vid': i, 'url': u, 'tags': cdata})
			
def connect2Browser(opts):
	logging.info('Попытка подключиться к браузеру 127.0.0.1:9222 ...')
	chrome_options = Options()
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222") # запустить предварительно Хром
	drv = webdriver.Chrome(opts.webdriver, options=chrome_options)
	time.sleep(P100ms*30) # пауза 3 при подключении к браузеру
	drv.set_page_load_timeout(30)
	drv.implicitly_wait(10)
	drv.maximize_window()
	discardChanges(drv,P100ms*20) #если при запуске есть не сохраненные изменения, то отменяем перед началом перехода по ссылке
	
	drv.get('https://youtube.com')
	try:
		testElm = WebDriverWait(drv, float(opts.timeout)).until(lambda x: x.find_element_by_id("avatar-btn"))
	except TimeoutException:
		logging.info("Выполните вход в аккаунт ютюб и перезапустите приложение.")
		exit(13)
	return drv

def getVID(strline):
	vid = '-'
	url = '-'
	if 'youtube.com' in strline:
		if strline.strip()[-5:]=='/edit':
			vid = strline.strip()[-16:-5]
		else:
			vid = strline.strip()[-11:]
		url='https://studio.youtube.com/video/{}/edit'.format(vid)
	return (vid, url)

def loadUrls(flname):
	res = {}
	with open(flname, 'r') as fl:
		for line in fl.readlines():
			if 'youtube.com' in line:
				if line.strip()[-5:]=='/edit':
					vid = line.strip()[-16:-5]
				else:
					vid = line.strip()[-11:]
				res[vid]='https://studio.youtube.com/video/{}/edit'.format(vid)
	return res

def tagsUpdate(drv, vid, url, o):
	svd = 0 # не привело к сохранению в видео ютюб
	drv.get(url)
	time.sleep(P100ms*20) # пауза 2
	try:
		testElm = WebDriverWait(drv, float(o.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
	except TimeoutException:
		logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
		
	# сформировать предлагаемый перечень тегов для замены
	ntags=getSEOtags(vid)

	# сохранить текущее состояние видео
	vidSEO1 = getYTseo4Vid(drv)
	logging.info('seo1: {}, real: {}, show: {}'.format(vidSEO1['seo'], vidSEO1['treal'], vidSEO1['tshow']))
	
	if len(ntags) > 0:
		logging.info('{} слов для формирования строки тегов'.format(len(ntags)))
		# подставить новую строку тегов
		tagStr = ''
		tagslen = 0
		for t in ntags:
			if tagslen < 500 and tagslen+len(t) < 500:
				tagStr += ','+t
				tagslen +=  len(t)
			else:
				break
		
		#logging.info('new tags:'+tagStr)

		inp = drv.find_element_by_id('text-input')
		clearAlltags(drv, inp)
		inp.click()
		pyperclip.copy(tagStr)
		time.sleep(PMIN)
		# вставить новую строку тегов
		#inp.send_keys(Keys.CONTROL, 'v')
		inp.send_keys(Keys.SHIFT, Keys.INSERT)
		time.sleep(P100ms*20)
		# проверить длинну <div slot="description" id="tags-count" class="style-scope ytcp-video-metadata-basics">155/500</div>
		tlen = 501
		while tlen > 500:
			# откорректировать длинну
			tagslength = drv.find_element_by_id('tags-count')
			#logging.info('tags-count: {}'.format(tagslength.text))
			chcnt = tagslength.text.split('/')
			#logging.info('tags-count: {}'.format(chcnt[0]))
			tlen = int(chcnt[0])
			if tlen > 500:
				delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
				if len(delbtns) > 0:
					delbtns[-1].click()
	else:
		logging.info('Нет слов для формирования строки тегов. Работаем с текущими')
		# вставлять нечего, далее работаем с тем, что есть

	# получить текущее состояние видео
	time.sleep(P100ms*10)
	vidSEO2 = getYTseo4Vid(drv)
	logging.info('seo2: {}, real2: {}, show2: {}'.format(vidSEO2['seo'], vidSEO2['treal'], vidSEO2['tshow']))
	
	#если рейтинг выше, то сохранить
	if (vidSEO2['treal'] > 49.99) or ((vidSEO2['treal'] > vidSEO1['treal']) and (vidSEO1['tshow']<0.1 or vidSEO2['tshow'] > vidSEO1['tshow'])):
		logging.info('!!! {:05.2f} >= {:05.2f} SAVE, seo2: {}, seo1: {}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['seo'], vidSEO1['seo'], vidSEO2['tshow'], vidSEO1['tshow']))
		# save this
		svd = 1
		elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
		if len(elms)>0:
			elms[0].click()
			time.sleep(PMIN)
	
	# продолжаем улучшать, пробуем удалять по одному тегу, возможно найдем более высокое значение рейтинга
	todo = True if vidSEO2['treal'] < 50 else False
	fup = 0
	fdown = 0
	iteration = 0
	vidSEOlast = vidSEO2.copy()
	while todo and iteration < 100 and fdown < 1 and fup < 2:
		iteration+=1
		tlen = readTagsLen(drv)
		deltrue = False
		deltag = ''
		if tlen > 0:
			#пытаемся удалить последний тег
			delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
			if len(delbtns) > 0:
				parentelm = delbtns[-1].find_element_by_xpath("./..")
				deltag = parentelm.get_attribute('vidiq-keyword')
				delbtns[-1].click()
				time.sleep(P100ms*20)
				deltrue = True

		vidSEOcur = getYTseo4Vid(drv)
		if deltrue:
			logging.info('удаляем тег: {}, seo: {}, real: {}, show: {}'.format(deltag,
						vidSEOcur['seo'], vidSEOcur['treal'], vidSEOcur['tshow']))

		if ( vidSEOlast['treal'] > vidSEOcur['treal'] ): #and ( abs(vidSEOlast['tshow'] - vidSEOcur['tshow']) <0.1 ):
			# уменьшился рейтинг после удаления тега - отменяем и завершаем подбор
			fdown +=1
			logging.info('--- {:05.2f} < {:05.2f} DISCARD'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal'])))
			vidSEOlast = vidSEOcur.copy()

		elif ( vidSEOlast['treal'] < vidSEOcur['treal']) and (vidSEO1['treal'] <= vidSEOcur['treal']):
			# сохранить и продолжить удалять до следующего изменения
			svd = 1
			fup +=1
			logging.info('!!! {:05.2f} > {:05.2f} SAVE {}'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal']), fup))
			elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
			if len(elms)>0:
				elms[0].click()
				time.sleep(PMIN)
			todo = True if vidSEOcur['treal'] < 50 else False
			vidSEOlast = vidSEOcur.copy()

	if fdown > 0:
		# cancel
		#logging.info(' :( discard, seo2 {:05.2f} =< {:05.2f} seo1.'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
		btn=drv.find_element_by_id('discard')
		btn.click()
		

	# сохранить рейтинг в БД
	time.sleep(P100ms*30) #пауза 3
	vidSEO2 = getYTseo4Vid(drv)
	if float(vidSEO2['treal']) > float(vidSEO1['treal']):
		logging.info('!!! текущий seo: {:05.2f} > старого {:05.2f}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['tshow'], vidSEO1['tshow']))
	elif float(vidSEO2['treal']) <= float(vidSEO1['treal']):
		logging.info('текущий seo: {:05.2f} <= старому {:05.2f}'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
	
	saveSEOupdate(vid, [vidSEO1, vidSEO2], svd)

def readTagsLen(drv):
	tagslength = drv.find_element_by_id('tags-count')
	chcnt = tagslength.text.split('/')
	return int(chcnt[0])

def readTagsCnt(drv):
	return len(drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']"))

def clearAlltags(drv, inp):
	# clear all tags
	# inp = drv.find_element_by_id('text-input')
	inp.click()
	time.sleep(PMIN)
	doit_clear=2
	while doit_clear>0:
		try:
			di = drv.find_element_by_id('clear-button')
			di.click()
			doit_clear=0
		except:
			logging.info("find_element_by_id('clear-button') Unexpected error: {}".format(sys.exc_info()[0]))
			inp.send_keys('wr1,wr2', Keys.ENTER)
			time.sleep(PMIN)
			doit_clear-=1
	pass

# удалить тег
def delTag(drv):
	try:
		btn = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
		if len(btn) > 1:
			di = drv.find_element_by_id('clear-button')
			di.click()
		else:
			btn[0].click()
	except:
		logging.info("ERROR: {}".format(sys.exc_info()[0]))


def saveindb(vdata):
	with orm.db_session:
		for t, d in vdata['tags'].items():
			dd = {'vid': vdata['vid'], 'url': vdata['url'], 'tag': t, 'seo': d['seo'], 'real': d['real'], 'tcount': d['tcount'], 'tpopular': d['tpopular'],
				  'tintitle': d['tintitle'], 'tindesc': d['tindesc'], 'triple': d['triple'], 'tshow': d['tshow'], 'ranked': d['ranked'], 'hivolume': d['hivolume']}
			
			ts = TagSEO.get(vid=vdata['vid'], url=vdata['url'], tag=t)
			if ts is not None:
				ts.dt=datetime.datetime.now()
				ts.seo = d['seo']
				ts.real = d['real']
				ts.tcount = d['tcount']
				ts.tpopular = d['tpopular']
				ts.tintitle = d['tintitle']
				ts.tindesc = d['tindesc']
				ts.triple = d['triple']
				ts.tshow = d['tshow']
				ts.ranked = d['ranked']
				ts.hivolume = d['hivolume']
				ts.data = dd
			else:
				ts = TagSEO(vid=vdata['vid'], url=vdata['url'], tag=t, seo=d['seo'],
					real=d['real'], tcount=d['tcount'], tpopular=d['tpopular'], tintitle=d['tintitle'], tindesc=d['tindesc'],
					triple=d['triple'], tshow=d['tshow'], ranked=d['ranked'], hivolume=d['hivolume'], data=dd)

def saveSEOupdate(vid, vdata, saved):
	with orm.db_session:
		if len(vdata)==2:
			TagUpdate(vid=vid, tags1=','.join(vdata[0]['tags']), real1=vdata[0]['treal'], tshow1=vdata[0]['tshow'], 
				tags2=','.join(vdata[1]['tags']), real2=vdata[1]['treal'], tshow2=vdata[1]['tshow'], jdata=vdata, saved=saved)
		else:
			logging.info('Error: vdata has to 2 elements!!!')
	pass

def getSEOtags(vid):
	tgs=[]
	with orm.db_session:
		for tg in TagSEO.select(lambda t: t.vid==vid).order_by(orm.raw_sql('seo desc, length(tag)')):
			#if tg.tag not in tgs:
			tgs+=[tg.tag]
	return tgs

def getDBTagSeo(vid, url, tag):
	resseo = 0
	with orm.db_session:
		ts = TagSEO.get(vid=vid, url=url, tag=tag)
		if ts is not None:
			resseo = ts.seo
	return resseo

def getExtags(minV, avgV, maxV):
	etags = []
	with orm.db_session:
		cnttags = orm.select(t.tag for t in TagSEO).count()
		etags = [tg for tg in orm.select(t.tag for t in TagSEO if orm.min(t.seo) < minV and orm.avg(t.seo) < avgV and orm.max(t.seo) < maxV)]
		logging.info('список слов исключений: {}/{}({:05.2f}%) шт. {}'.format(len(etags), cnttags, float(len(etags)*100.0/cnttags), etags))
	return etags

def getDBtags(vid, url, extags=[]):
	restags = []
	with orm.db_session:
		# теги видео
		tags=[t.tag for t in TagSEO.select(lambda t: t.url==url and t.vid==vid)]
		# теги исключения
		# extags
		if extags is None:
			extags = []
		#logging.info(tags)
		for tss in TagSEO.select(lambda ts: ts.vid!=vid and ts.url!=url).order_by(orm.desc(TagSEO.seo)):
			if tss.tag not in tags and tss.tag not in extags:
				restags+=[tss.tag]
		#logging.info(restags)
	return restags

def getTitleTags(vid, url, title, addtags=[]):
	restags = []
	ats = []
	loctitle = title.lower()
	with orm.db_session:
		# теги видео
		tags = [t.tag for t in TagSEO.select(lambda t: t.url==url and t.vid==vid)]
		for t in addtags:
			if t not in tags:
				ats.append(t)
		for tss in TagSEO.select(lambda ts: ts.vid!=vid and ts.url!=url and ts.tag in loctitle).order_by(orm.desc(TagSEO.seo)):
			if tss.tag not in tags and tss.tag not in addtags:
				restags+=[tss.tag]
		#logging.info(restags)
		for t in ats:
			if t not in restags:
				restags.append(t)
	return restags

def getScores(dr, tdata):
	#get h1 stat-value-undefined two times
	elms=dr.find_elements_by_class_name('stat-value-undefined')
	elm1 = elms[0].find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tdata['real'] = float(elm1.text)

	elm1 = elms[1].find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tdata['tshow'] = float(elm1.text)

	#get stat-value-tag-count
	elm=dr.find_element_by_class_name('stat-value-tag-count')
	elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tdata['tcount'] = int(elm1.text)

	#get stat-value-tag-volume
	elm=dr.find_element_by_class_name('stat-value-tag-volume')
	elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tdata['tpopular'] = int(elm1.text)

	#get stat-value-keywords-in-title
	elm=dr.find_element_by_class_name('stat-value-keywords-in-title')
	elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tdata['tintitle'] = int(elm1.text)

	#get stat-value-keywords-in-description
	try:
		elm=dr.find_element_by_class_name('stat-value-keywords-in-description')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		tdata['tindesc'] = int(elm1.text)
	except:
		logging.info("stat-value-keywords-in-description Unexpected error: {}".format(sys.exc_info()[0]))

	#get stat-value-tripled-keywords
	try:
		elm=dr.find_element_by_class_name('stat-value-tripled-keywords')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		tdata['triple'] = int(elm1.text)
	except:
		logging.info("stat-value-tripled-keywords Unexpected error: {}".format(sys.exc_info()[0]))

				
	#get <span class="stat-value stat-value-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
	try:
		elm=dr.find_element_by_class_name('stat-value-ranked-tags')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		tdata['ranked'] = int(elm1.text)
	except:
		logging.info("stat-value-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

	#get <span class="stat-value stat-value-high-volume-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
	try:
		elm=dr.find_element_by_class_name('stat-value-high-volume-ranked-tags')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		tdata['hivolume'] = float(elm1.text)
	except:
		logging.info("stat-value-high-volume-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

	pass

# --update=3 собрать теги и рейтинги
def getYTseo4Vid(drv):
	tags=[]
	for elm in drv.find_elements_by_xpath("//*[@id='child-input']//ytcp-chip[@role='button']"):
		text = elm.get_attribute('vidiq-keyword')
		tags+=[text]
	
	#get SEO score
	seo=-1
	elm=drv.find_element_by_class_name('stat-value-seo-score')
	try:
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		seo = float(elm1.text)
	except:
		logging.info("stat-value-seo-score Unexpected error: {}".format(sys.exc_info()[0]))

	# собрать рейтинги
	elms=drv.find_elements_by_class_name('stat-value-undefined')
	elm1 = elms[0].find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tr = float(elm1.text)

	elm1 = elms[1].find_elements_by_xpath(".//span[@class='value-inner']")[0]
	tsh = float(elm1.text)

	try:
		elm=drv.find_element_by_class_name('stat-value-ranked-tags')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		tranked = int(elm1.text)
	except:
		logging.info("stat-value-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

	#get <span class="stat-value stat-value-high-volume-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
	try:
		elm=drv.find_element_by_class_name('stat-value-high-volume-ranked-tags')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		thivolume = float(elm1.text)
	except:
		logging.info("stat-value-high-volume-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))


	return {'tags': tags, 'treal': tr, 'tshow': tsh, 'tranked': tranked, 'thivolume': thivolume, 'seo': seo}

def CheckSavedData(vid, dt):
	cdjson = None
	with orm.db_session:
		datas = TagUpdate.select(lambda t: t.vid==vid and  t.dt > dt and t.saved==1).order_by(orm.desc(TagUpdate.dt))
		for d in datas:
			cdjson={'dt': d.dt, 'rate': d.jdata, }
			break
	return cdjson

def discardChanges(drv, wt=0.5):
	#logging.info('cancel changes')
	try:
		btn=drv.find_element_by_id('discard')
		btn.click()
		time.sleep(wt)
	except:
		pass

def rateWords(drv, opts):
	words = []
	addwrds = []
	with open(opts.infile, 'r', encoding='utf-8') as fl:
		linecnt = 0
		for line in fl.readlines():
			linecnt+=1
			if linecnt==1:
				url = line.strip()
			else:
				t = line.strip()
				if len(t)>0:
					words+=[t]
	wordscnt= len(words)				
	logging.info('Прочитано из файла слов: {}'.format(wordscnt))
	logging.info('Перехожу по ссылке: {}'.format(url))
	
	drv.get(url)
	time.sleep(P100ms*30)  # пауза 3
	try:
		testElm = WebDriverWait(drv, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
	except TimeoutException:
		logging.info("Превышено время ожидания загрузки страницы.")
		exit(11)
		
	inp = drv.find_element_by_id('text-input')
	clearAlltags(drv, inp)
	inp.click()
	curw=0
	for w in words:
		curw+=1
		inp.click()
		if opts.clipboard=='1':
			pyperclip.copy(w)
			time.sleep(PMIN)
			inp.send_keys(Keys.SHIFT, Keys.INSERT)
		else:
			inp.send_keys(w+',',Keys.ENTER)
		time.sleep(P100ms*30) # пауза 3
		# подождать появления подсказок
		sugText=''
		try:
			wElm = WebDriverWait(drv, float(2)).until(lambda x: x.find_element_by_class_name("vidiq-studio-beta-keyword-text"))
			sugText=wElm.text
		except TimeoutException:
			pass
		if len(sugText)>0:
			restxt = []
			for el in drv.find_elements_by_class_name("vidiq-studio-beta-keyword-text"):
				if el.is_displayed():
					restxt += [el.text]
			if len(restxt) > 0:
				sugText=';'.join(restxt)
				addwrds+=restxt

		#get SEO score
		elm=drv.find_element_by_class_name('stat-value-seo-score')
		elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
		logging.info('{:03d}/{:03d} ;{:05.2f};{};{}'.format(curw, wordscnt, float(elm1.text), w, sugText))
		delTag(drv)
	discardChanges(drv, PMIN)
	newresfile = '!addwords_{}.txt'.format(datetime.datetime.now().strftime('%y%m%d_%H%M'))
	with open(newresfile, 'w', encoding='utf-8') as flres:
		for adw in addwrds:
			flres.writelines([adw+'\n'])
	exit(999)


def getVideoList(opts):
	# наименование сервиса Google
	api_service_name = "youtube"
	api_version = "v3"
	youtube = googleapiclient.discovery.build(api_service_name, api_version, developerKey=opts.apiKey, cache_discovery=False)
	
	playlists=[]
	npage = 1
	tstr = ''
	# step 2 - получить список плейлистов
	######################################
	logging.info("Получаем список плейлистов канала")
	request = youtube.playlists().list(part="id", maxResults=10, channelId=opts.chID) 
	response = request.execute()
	playlists.extend(response["items"])
	
	while 'nextPageToken' in response and response["nextPageToken"]:
		request = youtube.playlists().list(part="id", maxResults=10, channelId=opts.chID, pageToken=response["nextPageToken"])
		response = request.execute()
		npage+=1
		playlists.extend(response["items"])
	#logging.info(playlists)
	logging.info("Получена информация о {} плейлистах канала".format(len(playlists)))
	plvideos = []
	npage = 0
	for pl in playlists:
		npage+=1
		logging.info("Запрос списка видео плейлиста {} из {}".format(npage, len(playlists)))
		request = youtube.playlistItems().list(
			part="snippet", 
			maxResults=50,
			playlistId=pl["id"] # плейлист
		)
		response = request.execute()
		totalResults = response["pageInfo"]["totalResults"]
		logging.info(f'Предстоит получить информацию о {totalResults} видео')
		for itm in response["items"]:
			#plvideos.extend([itm["snippet"]])
			req = youtube.videos().list(part="snippet", maxResults=1, id=itm["snippet"]["resourceId"]["videoId"])
			resp = req.execute()
			for elm in resp["items"]:
				plvideos.extend([elm])
			
		while 'nextPageToken' in response:# and response["nextPageToken"] is not None:
			npToken = response.get('nextPageToken','not Found')
			
			request = youtube.playlistItems().list(
				part="snippet", # заменить на snippet 
				maxResults=50,
				pageToken=npToken,
				playlistId=pl["id"] # плейлист
			)
			response = request.execute()
			#logging.info(response['items'])
			for itm in response["items"]:
				#plvideos.extend([itm["snippet"]]) # накапливаем список видео, обновляем характеристики видео, далее собираем только статистику
				req = youtube.videos().list(part="snippet", maxResults=1, id=itm["snippet"]["resourceId"]["videoId"])
				resp = req.execute()
				for elm in resp["items"]:
					plvideos.extend([elm])

	#plvideos["channelId"]
	#plvideos["resourceId"]["videoId"]
	yturls=[]
	if len(opts.infile)>1:
		with open(opts.infile, 'r') as infl:
			for line in infl.readlines():
				if 'youtube.com' in line:
					yturls.append(line.strip())
	
	res = getFilename(opts)
	if opts.xls == '1':
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Видео"
		cr = 1
		ws.cell(row=cr, column=1, value='channel')
		ws.cell(row=cr, column=2, value='url')
		ws.cell(row=cr, column=3, value='short description')
		ws.cell(row=cr, column=4, value='zero time exists')
		ws.cell(row=cr, column=5, value='zero time hashtag exists')
		ws.cell(row=cr, column=6, value='title')
		ws.cell(row=cr, column=7, value='description')


		for v in plvideos:
			cr+=1
			ws.cell(row=cr, column=1, value=v["snippet"]["channelTitle"])
			ws.cell(row=cr, column=2, value='https://youtube.com/watch?v={}'.format(v["id"]))
			ws.cell(row=cr, column=3, value=shortdescription(v["snippet"]["title"], v["snippet"]["description"], int(opts.short)))
			ws.cell(row=cr, column=4, value=zerotime_exists(v["snippet"]["description"]))
			ws.cell(row=cr, column=5, value=zerotimehashtag_exists(v["snippet"]["description"]))
			ws.cell(row=cr, column=6, value=v["snippet"]["title"])
			ws.cell(row=cr, column=7, value=v["snippet"]["description"])
			

		wb.save(res)
	else:
		with open(res, 'w', encoding='utf-8') as fl:
			# if opts.chvideos=='1':
			# 	for v in plvideos:
			# 		if v["channelId"] == opts.chID:
			# 			fl.writelines(['{};https://youtube.com/watch?v={}\n'.format(v["channelId"], v["resourceId"]["videoId"])])
			# else:
			
			if len(yturls) > 0:
				for v in plvideos:
					url = 'https://youtube.com/watch?v={}'.format(v["id"])
					if url not in yturls:
						fl.writelines([url+'\n'])
			else:
				fl.writelines(['https://youtube.com/watch?v={}\n'.format(v["id"]) for v in plvideos])
	return res

def getFilename(opts):
	if opts.xls == '1':
		res='{}_{}.xlsx'.format(opts.outflname, datetime.datetime.now().strftime('%y%m%d_%H%M'))
	else:
		res='{}_{}.txt'.format(opts.outflname, datetime.datetime.now().strftime('%y%m%d_%H%M'))
	return res

def shortdescription(t, d, slen=100):
	res = False
	if (t == d) or (len(d) < slen):
		res = True
	return res

def zerotime_exists(d):
	res = False
	if '00:00' in d:
		res = True
	return res
def zerotimehashtag_exists(d):
	res = False
	if ('00:00' in d) and ('#' in d):
		res = True
	return res
#------------------------------------- from check_yt_videos.py --------------------
def tags_openxls(flname):	
	logging.info(f'Чтение списка ссылок из файла {flname}')
	#logging.info(os.path.dirname(opts.xls))
	wb = openpyxl.load_workbook(filename=flname, read_only=False, keep_vba=True)
	if 'Обновление' in wb.sheetnames:
		wss = wb['Обновление']
	else:
		logging.info('source worksheet did not find')
		exit(2)
	if 'Статистика' in wb.sheetnames:
		wsd = wb['Статистика']
	else:
		logging.info('destination worksheet did not find')
		exit(3)
	# if 'Исключения' in wb.sheetnames:
	# 	wse = wb['Исключения']
	# else:
	# 	logging.info('exclusions worksheet did not find')
	if 'Теги' in wb.sheetnames:
		wst= wb['Теги']
	else:
		logging.info('tags worksheet did not find')

	#newWBfile = '{}\\Ссылки_{}.xlsm'.format(os.path.dirname(flname), datetime.datetime.now().strftime('%y%m%d_%H%M'))
	newWBfile = 'Ссылки_{}.xlsm'.format(datetime.datetime.now().strftime('%y%m%d_%H%M'))
	#logging.info(newWBfile)
	#return (wb, wss, wsd, wse, wst, newWBfile)
	return (wb, wss, wsd, wst, newWBfile)
	
	
def loadUrlsFromTxt(flname):
	res = []
	if os.path.exists(flname):
		indx = 0
		with open(flname, 'r', encoding='utf-8') as fl:
			for l in fl.readlines():
				curval = l.strip()
				if ('studio.youtube.com' in curval) :
					url = curval.strip()
				elif ('youtube.com' in curval) :
					url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/edit'
				res.append({'url': url, 'num': indx, 'title': 'Видео из текстового файла'})
	return res

def loadVideoUrlsFromTxt(flname):
	res = []
	if os.path.exists(flname):
		indx = 0
		with open(flname, 'r', encoding='utf-8') as fl:
			for l in fl.readlines():
				url = l.strip()
				indx+=1
				res.append({'url': url, 'num': indx,})
	return res

def check_list(opts):
	started_at = time.monotonic()
	#if opts.tags == '0':  обновление статистики чек-листа
	#(wb, wss, wsd, wse, wst, newWBfile) = tags_openxls(opts.infile)
	isxls = True if '.xls' in opts.infile else False
	if isxls:
		(wb, wss, wsd, wst, newWBfile) = tags_openxls(opts.infile)
		tags_clearStatSheet(opts, wsd)
		urls = tags_getUrls(wss, opts.owner)
	else:
		urls = loadUrlsFromTxt(opts.infile)
	urlslen = len(urls)
	logging.info(f'---------- Очередь для обработки {urlslen} ссылок')

	odt = datetime.datetime.strptime(opts.dt, '%Y-%m-%d %H:%M')
	logging.info('---- дата свежести сохранненых данных {}'.format(odt))
	if urlslen > 0:
		driver = connect2Browser(opts)
	crow=1
	for u in urls:
		resttime = (urlslen - crow) * (time.monotonic() - started_at) / crow
		if isxls:
			logging.info('-------- Обработка {:03d} из {:03d} ({:05.2f} %), осталось {}, {}'.format(crow, urlslen, 100.00*crow/urlslen, 
					time.strftime('%H:%M:%S', time.gmtime(resttime)), u['title']))
		else:
			logging.info('-------- Обработка {:03d} из {:03d} ({:05.2f} %), осталось {}, {}'.format(crow, urlslen, 100.00*crow/urlslen, 
					time.strftime('%H:%M:%S', time.gmtime(resttime)), u['url']))
		crow+=1
		svdata = tags_CheckSavedData(u['url'], odt)
		if svdata is not None and float(opts.seo) <= float(svdata["seo"]):
			if isxls:
				# записать в эксель данные из БД
				tags_write2xls(wsd, crow, svdata)
			else:
				logging.info('в БД seo = {}'.format(svdata["seo"]))
		else:
			#запустить сбор данных и сохранение для дальнейшего использования
			if svdata is not None:
				logging.info('в БД seo = {}'.format(svdata["seo"]))
			else:
				logging.info('сбор информации о видео')
			driver.get(u['url'])
			if isxls:
				wsd.cell(row=crow, column=1, value=u['num'])
				wsd.cell(row=crow, column=2, value=u['title'])
				wsd.cell(row=crow, column=3, value=u['url'])
			time.sleep(P100ms*10)
			# Обрабатывать ошибку "чужое видео" пример https://studio.youtube.com/video/ez5hxF1jP-A/edit
			try:
				myElem = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
			except TimeoutException:
				if driver.find_element_by_id("error-image"):
					logging.info("Видео не доступно или показатели seo не успели отобразиться на веб-странице. {}, url={}".format(u['title'], u['url']))
					continue
				logging.info("Loading took too much time!")

			driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
			time.sleep(PMIN)

			cdata={'num': u['num'],
				   'title': u['title'],
				   'url': u['url'],}
			if not isxls:
				# read title from web page
				inpts = driver.find_elements_by_xpath("//ytcp-mention-input")
				vid_title = inpts[0].text
				cdata['title'] = vid_title
			
			#get SEO score
			elm=driver.find_element_by_class_name('stat-value-seo-score')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			logging.info('seo {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=4, value=float(elm1.text))
			cdata['seo'] = elm1.text
			
			#get h1 stat-value-undefined two times
			elms=driver.find_elements_by_class_name('stat-value-undefined')
			#logging.info('elms count {}'.format(len(elms)))
			elm1 = elms[0].find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('undefined {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=5, value=float(elm1.text))
			cdata['stat-value1'] = elm1.text

			elm1 = elms[1].find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('undefined {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=11, value=float(elm1.text))
			cdata['stat-value2'] = elm1.text

			#get stat-value-tag-count
			elm=driver.find_element_by_class_name('stat-value-tag-count')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('count {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=6, value=int(elm1.text))
			cdata['tag-count'] = elm1.text

			#get stat-value-tag-volume
			elm=driver.find_element_by_class_name('stat-value-tag-volume')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('volume {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=7, value=int(elm1.text))
			cdata['tag-volume'] = elm1.text

			#get stat-value-keywords-in-title
			elm=driver.find_element_by_class_name('stat-value-keywords-in-title')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('title {}'.format(elm1.text))
			if isxls:
				wsd.cell(row=crow, column=8, value=int(elm1.text))
			cdata['keywords-in-title'] = elm1.text

			#get stat-value-keywords-in-description
			try:
				elm=driver.find_element_by_class_name('stat-value-keywords-in-description')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('seo {}'.format(elm1.text))
				if isxls:
					wsd.cell(row=crow, column=9, value=int(elm1.text))
				cdata['keywords-in-description'] = elm1.text
			except:
				logging.info("stat-value-keywords-in-description Unexpected error: {}".format(sys.exc_info()[0]))

			#get stat-value-tripled-keywords
			try:
				elm=driver.find_element_by_class_name('stat-value-tripled-keywords')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('tripled-keywords {}'.format(elm1.text))
				if isxls:
					wsd.cell(row=crow, column=10, value=int(elm1.text))
				cdata['tripled-keywords'] = elm1.text
			except:
				logging.info("stat-value-tripled-keywords Unexpected error: {}".format(sys.exc_info()[0]))

			
			#get <span class="stat-value stat-value-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
			try:
				elm=driver.find_element_by_class_name('stat-value-ranked-tags')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('undefined {}'.format(elm1.text))
				if isxls:
					wsd.cell(row=crow, column=12, value=int(elm1.text))
				cdata['ranked-tags'] = elm1.text
			except:
				logging.info("stat-value-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

			#get <span class="stat-value stat-value-high-volume-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
			try:
				elm=driver.find_element_by_class_name('stat-value-high-volume-ranked-tags')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('undefined {}'.format(elm1.text))
				if isxls:
					wsd.cell(row=crow, column=13, value=float(elm1.text))
				cdata['volume-ranked-tags'] = elm1.text
			except:
				logging.info("stat-value-high-volume-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

			#get checklist
			try:
				driver.find_element_by_class_name('stat-box-checklist').click()
				elm=driver.find_element_by_class_name('stat-value-checklist')
				#logging.info(elm.get_attribute('innerHTML'))
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				logging.info('check-list {}'.format(elm1.text))
				col=14
				if isxls:
					wsd.cell(row=crow, column=col, value=int(elm1.text))
				cdata['chl-14'] = int(elm1.text)

				elm = driver.find_element_by_class_name('seo-checklist')
				for el in elm.find_elements_by_xpath(".//li")[:9]:
					col+=1
					if isxls:
						wsd.cell(row=crow, column=col, value=1 if 'checked' == el.get_attribute('class') else 0)
					cdata['chl-{}'.format(col)] = 1 if 'checked' == el.get_attribute('class') else 0
			except:
				logging.info("stat-box-checklist Unexpected error: {}".format(sys.exc_info()[0]))

			tags_SaveCheckData(cdata)
		# except:
		# 	continue
	if isxls:
		wb.save(newWBfile)
		logging.info(f'Результаты сохранены в файле {newWBfile}')
	logging.info('Работа сценария завершена за {:.3f} сек.'.format(time.monotonic() -started_at))
	return

def set_tags(opts):
	started_at = time.monotonic()
	#if opts.tags == '1': #подставляем теги
	#(wb, wss, wsd, wse, wst, newWBfile) = tags_openxls(opts.infile)
	(wb, wss, wsd, wst, newWBfile) = tags_openxls(opts.infile)

	print('------------------tags')
	urltags = tags_getUrlsTags(wss, opts.owner, wst)
	print('------------------urls count ', len(urltags))
	#logging.info('key:{}, title:{}, tags:{}'.format(urltags[0]['key'], urltags[0]['title'], urltags[0]['tags']))
	crow=1
	urlslen = len(urltags)
	if urlslen > 0:
		driver = connect2Browser(opts)
	for u in urltags:
		logging.info('-------- Обработка {:03d} из {:03d} ({:06.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen,u['title']))
		crow+=1

		svdata = None #CheckSavedData(u['url'], odt)
		if svdata is not None:
			# записать в эксель данные из БД
			#write2xls(wsd, crow, svdata)
			pass
		else:
			#запустить установку тегов
			driver.get(u['url'])
			time.sleep(P100ms*10)
			# Обрабатывать ошибку "чужое видео" пример https://studio.youtube.com/video/ez5hxF1jP-A/edit
			try:
				myElem = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
			except TimeoutException:
				if driver.find_element_by_id("error-image"):
					logging.info("Видео не доступно. {}, url={}".format(u['title'], u['url']))
					continue
				logging.info("Loading took too much time!")

			driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
			
			cdata={'num': u['num'],
				   'title': u['title'],
				   'url': u['url'],}

			#get SEO1 score
			elm=driver.find_element_by_class_name('stat-value-seo-score')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			logging.info('seo1 {}'.format(elm1.text))
			cdata['seo1'] = float(elm1.text)
			
			# Обновлять только если текущее значение seo меньше заданного граничного из эксель
			#logging.info('seo {} vs rate {}'.format(cdata['seo1'], float(u['taginfo']['maxseo'])))
			if cdata['seo1'] < float(u['taginfo']['maxseo']):

				elq = driver.find_element_by_id('text-input')
				elq.click()
				pyperclip.copy('word1, word2')
				time.sleep(PMIN)
				elq.send_keys(Keys.SHIFT, Keys.INSERT)
				time.sleep(PMIN)
				#clear all tags
				try:
					di = driver.find_element_by_id('clear-button')
					di.click()
				except:
					logging.info("find_element_by_id('clear-button') Unexpected error: {}".format(sys.exc_info()[0]))

				#add new tags
				elq.click()
				pyperclip.copy(u['taginfo']['newtags'])
				time.sleep(PMIN)
				elq.send_keys(Keys.SHIFT, Keys.INSERT)
				time.sleep(P100ms*30) # пауза 3

				#get SEO score
				elm=driver.find_element_by_class_name('stat-value-seo-score')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				logging.info('seo2 {}'.format(elm1.text))
				cdata['seo2'] = float(elm1.text)
				
				if cdata['seo2'] < 0.01:
					# если не успело значение обновиться, то подождем еще и вычитаем повторно
					time.sleep(P100ms*20) # пауза 2
					#get SEO score
					elm=driver.find_element_by_class_name('stat-value-seo-score')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					logging.info('seo2 {}'.format(elm1.text))
					cdata['seo2'] = float(elm1.text)

				if cdata['seo2'] > cdata['seo1']:
					# примем решение сохранить если значение увеличилось
					logging.info('save')
					elms=driver.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
					if len(elms)>0:
						elms[0].click()
						time.sleep(P100ms*10)
				else:
					logging.info('cancel')
					elm=driver.find_element_by_id('discard')
					elm.click()
					time.sleep(PMIN)
			else:
				logging.info('seo {} > tag rate {}'.format(cdata['seo1'], float(u['taginfo']['maxseo'])))	
			
			#exit(998) #удалить перед релизом, только для теста
	logging.info('Работа сценария завершена за {:.3f} сек.'.format(time.monotonic() - started_at))
	return

def tags_getUrls(ws, owner): #wse):
	#exurls = []
	urls=[]
	# doit = True
	# r = 1
	# emptyrows = 0
	# # сформировать список исключений
	# while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
	# 	r+=1
	# 	curval = wse.cell(row=r, column=6).value
	# 	if curval is not None:
	# 		if 'youtube.com' in curval:
	# 			exurls+=[curval]
	# 	else:
	# 		emptyrows+=1
	# 	if emptyrows > 3:
	# 		doit = False
	
	# прочитать ссылки на ютюб для сбора информации по чек-лист
	doit = True
	r = 1
	emptyrows = 0
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = ws.cell(row=r, column=6).value # link
		curown = ws.cell(row=r, column=4).value # owner
		if curval is not None:
			if ('youtube.com' in curval) and (curown == owner): #(not curval in exurls):
				url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/edit'
				logging.info(url)
				urls+=[{'url': url, 'num': ws.cell(row=r, column=1).value, 'title': ws.cell(row=r, column=5).value}]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	return urls

def tags_getUrlsTags(ws, owner, wst): #wse, 
	titlemask = {} # маски названий
	#exurls = [] # исключения
	urls=[] # видео

	# doit = True
	# r = 1
	# emptyrows = 0
	# # сформировать список исключений
	# while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
	# 	r+=1
	# 	curval = wse.cell(row=r, column=6).value
	# 	if curval is not None:
	# 		if 'youtube.com' in curval:
	# 			exurls+=[curval]
	# 	else:
	# 		emptyrows+=1
	# 	if emptyrows > 3:
	# 		doit = False
	
	
	doit = True
	r = 1
	emptyrows = 0
	# сформировать список масок названий
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = wst.cell(row=r, column=2).value
		if curval is not None:
			if '%' in curval:
				curkey = curval.replace('%', '')
				titlemask[curkey]={'newtags': wst.cell(row=r, column=4).value, 'maxseo': wst.cell(row=r, column=3).value, }
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	#logging.info(titlemask.keys())
	# прочитать ссылки на ютюб для установки тегов
	doit = True
	r = 1
	emptyrows = 0
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = ws.cell(row=r, column=6).value
		title = ws.cell(row=r, column=5).value
		curown = ws.cell(row=r, column=4).value
		if curval is not None and title is not None:
			if ('youtube.com' in curval) and (curown == owner): # (not curval in exurls):
				for k in titlemask.keys():
					if k in title:
						url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/edit'
						logging.info(url)
						urls+=[{'url': url, 'num': ws.cell(row=r, column=1).value, 'title': ws.cell(row=r, column=5).value, 'key': k, 'taginfo': titlemask[k], }]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False

	return urls

def anl_getVUrls(ws, owner):  
	urls=[] # видео
	# прочитать ссылки на ютюб для установки тегов
	doit = True
	r = 1
	emptyrows = 0
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = ws.cell(row=r, column=6).value
		title = ws.cell(row=r, column=5).value
		curown = ws.cell(row=r, column=4).value
		if curval is not None and title is not None:
			if ('youtube.com' in curval) and (curown == owner): # (not curval in exurls):
				#logging.info(curval)
				vID = curval.split('=')[-1]
				#logging.info(vID)
				url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/analytics/tab-overview/period-lifetime'
				#https://studio.youtube.com/video/i09vBTTHYOg/analytics/tab-overview/period-since_publish
				#logging.info(url)
				urls+=[{'url': url, 'num': ws.cell(row=r, column=1).value, 'title': ws.cell(row=r, column=5).value, 'videoId': vID}]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	return urls

def getVidTags4Import(indt):
	viddata = {}
	cnt_tags = 0
	max_len = 0
	lasturl = '-'
	vid = '-'
	url = '-'
	odt = datetime.datetime.strptime(indt, '%Y-%m-%d %H:%M')
	with orm.db_session:
		curvidTags = []
		#vidtags = TagImport.select().order_by(TagImport.url)
		vidurls = orm.select(t.url for t in TagImport).distinct()
		lenvid = len(vidurls) # заменить на кол-во видео, а не тегов
		vidtags = orm.select((t.url, t.tag) for t in TagImport).distinct().order_by(orm.raw_sql("url"))
		avgTags = {}
		curindx = 0
		for vt in vidtags:
			
			curl = vt[0]
			tag = vt[1]
			if tag not in avgTags.keys():
				tag_seo = -1
				# запросить в БД
				for ts in orm.select((ts.tag, orm.raw_sql('avg(seo)')) for ts in TagSEO if ts.tag == tag):
					if ts[1] is not None:
						tag_seo = ts[1]
					else:
						tag_seo = -1
				avgTags[tag] = tag_seo
			else:
				# прочитать сохраненное значение
				tag_seo = avgTags[tag]

			if curl == lasturl:
				if tag not in curvidTags:
					# tag_seo = -1
					# for ts in orm.select((ts.tag, orm.raw_sql('avg(seo)')) for ts in TagSEO if ts.tag == tag):
					# 	tag_seo = ts[1]
					viddata[vid]['tags'].append({'t': tag, 'ttype': '?', 'seoavg': tag_seo})
					cnt_tags+=1
					if max_len < len(tag):
						max_len = len(tag)
			else:
				curindx += 1
				# отсортировать теги предыдущего видео
				if vid != '-' and len(viddata[vid]['tags']) > 0:
					viddata[vid]['tags'] = sorted(viddata[vid]['tags'], key=lambda k: k['seoavg'], reverse=True)
				(vid, url) = getVID(curl)
				# получить текущий список тегов для видео
				curvidTags = [ts.tag for ts in TagSEO.select(lambda x: x.url==curl and x.dt > odt).order_by(TagSEO.tag)]
				if tag not in curvidTags:
					# tag_seo = -1
					# for ts in orm.select((ts.tag, orm.raw_sql('avg(seo)')) for ts in TagSEO if ts.tag == tag):
					# 	tag_seo = ts[1]
					viddata[vid]={'url': url, 'tags': [{'t': tag, 'ttype': '?', 'seoavg': tag_seo}]}
					cnt_tags+=1
					if max_len < len(tag):
						max_len = len(tag)
				else:
					viddata[vid]={'url': url, 'tags': []}
				lasturl = curl
				logging.info(f'{curindx:03d}/{lenvid:03d} - подготовка тегов для видео и сортировка')
		if vid != '-' and len(viddata[vid]['tags']) > 0:
			viddata[vid]['tags'] = sorted(viddata[vid]['tags'], key=lambda k: k['seoavg'], reverse=True)
	return (viddata, cnt_tags, max_len)

def getNewTags4Vid1(curl, alltags, indt):
	odt = datetime.datetime.strptime(indt, '%Y-%m-%d %H:%M')
	restags = []
	with orm.db_session:
		vidTags = [ts.tag for ts in TagSEO.select(lambda x: x.url==curl and x.dt > odt).order_by(TagSEO.tag)]
	for t in alltags:
		if t not in vidTags:
			restags.append(t)
	return restags

def getNewTags4Vid2(curl, alltags, indt):
	odt = datetime.datetime.strptime(indt, '%Y-%m-%d %H:%M')
	restags = []
	with orm.db_session:
		vidTags = [ts.tag for ts in TagSEO.select(lambda x: x.url==curl and x.dt > odt).order_by(TagSEO.tag)]
	for t in alltags:
		if t['t'] not in vidTags:
			restags.append(t)
	return restags

def tags_CheckSavedData(url, dt):
	cdjson = None
	with orm.db_session:
		cd = CheckData.get(url=url)
		if cd is not None:
			if cd.dt > dt:
				cdjson = cd.data
	return cdjson

def tags_SaveCheckData(cdata):
	with orm.db_session:
		cd = CheckData.get(url=cdata['url'])
		if cd is None:
			cd = CheckData(url=cdata['url'], data=cdata)
		else:
			cd.data=cdata
			cd.dt=datetime.datetime.now()

def tags_write2xls(ws, crow, data):
	inum = ''
	if 'num' in data:
		if data["num"] is not None:
			inum = int(data["num"])
	ws.cell(row=crow, column=1, value=inum)
	ws.cell(row=crow, column=2, value=data["title"] if 'title' in data else '')
	ws.cell(row=crow, column=3, value=data["url"] if 'url' in data else '')

	ws.cell(row=crow, column=4, value=float(data["seo"]) if 'seo' in data else '')

	ws.cell(row=crow, column=5, value=float(data["stat-value1"]) if 'stat-value1' in data else '')
	ws.cell(row=crow, column=6, value=int(data["tag-count"]) if 'tag-count' in data else '')
	ws.cell(row=crow, column=7, value=int(data["tag-volume"]) if 'tag-volume' in data else '')
	ws.cell(row=crow, column=8, value=int(data["keywords-in-title"]) if 'keywords-in-title' in data else '')
	ws.cell(row=crow, column=9, value=int(data["keywords-in-description"]) if 'keywords-in-description' in data else '')
	ws.cell(row=crow, column=10, value=int(data["tripled-keywords"]) if 'tripled-keywords' in data else '')
	ws.cell(row=crow, column=11, value=float(data["stat-value2"]) if 'stat-value2' in data else '')
	ws.cell(row=crow, column=12, value=float(data["ranked-tags"]) if 'ranked-tags' in data else '')
	ws.cell(row=crow, column=13, value=float(data['volume-ranked-tags']) if 'volume-ranked-tags' in data else '')
	
	for i in range(14,24):
		key = 'chl-{}'.format(i)
		ws.cell(row=crow, column=i, value=int(data[key]) if key in data else '')

def tags_clearStatSheet(opts, wsd):
	crow = 1
	emptyrows = 0
	todo = True
	while todo and emptyrows < 3:
		crow+=1
		if not wsd.cell(row=crow, column=4).value is None:
			for n in range(1,23):
				wsd.cell(row=crow, column=n).value = None
		else:
			emptyrows+=1

def loadTags(flname):
	restags=[]
	if os.path.exists(flname):
		with open(flname, 'r', encoding='utf-8') as fl:
			restags = [l.strip() for l in fl.readlines()]
	return restags

def addAndEstimate(opts):
	logging.info(f'Загрузка списка видео из файла {opts.infile}')
	urls=loadUrls(opts.infile)
	urlslen = len(urls.keys())
	logging.info('Подготовлено {} видео для анализа'.format(urlslen))
	tagsIN=loadTags(opts.addtags)
	tagslen = len(tagsIN)
	logging.info('Подготовлено {} тегов для анализа'.format(tagslen))
	
	driver = connect2Browser(opts)
	indx = 0
	# проход по видео
	for vid, cururl in urls.items():
		indx+=1
		exitOnKey()
		logging.info(f'{indx:03d}/{urlslen:03d} открытие страницы {vid} видео {cururl}')
		checkPauseKey() # key p
		cdata = {}
		tags = getNewTags4Vid1(cururl, tagsIN, opts.dt)
		tagslen = len(tags)
		logging.info(f'{tagslen} тегов для оценки под текущим видео с учетом даты {opts.dt}')
		
		if tagslen > 0:
			driver.get(cururl)
			time.sleep(P100ms*20) # пауза 2
		
			try:
				testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
			except TimeoutException:
				logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
				continue
		else:
			continue			
		inpts = driver.find_elements_by_xpath("//ytcp-mention-input")
		vid_title = inpts[0].text
		logging.info(f'Наименование видео {vid_title}')
		# добавить два слова для очистки тегов, чтобы гарантированно отображалась кнопка удалить все теги
		elq = driver.find_element_by_id('text-input')
		elq.click()
		# #clear all tags
		doit_clear=2
		while doit_clear>0:
			try:
				di = driver.find_element_by_id('clear-button')
				di.click()
				doit_clear=0
			except:
				logging.info("find_element_by_id('clear-button') Unexpected error: {}, {}, {}".format(sys.exc_info()[0], sys.exc_info()[1], traceback.format_exc()))
				elq.send_keys('wr1,wr2', Keys.ENTER)
				time.sleep(PMIN)
				doit_clear-=1

		# по тегам сохранить рейтинг
		tagscnt=tagslen
		crow = 0
		times = []
		for t in tags:
			checkPauseKey() # key p
			started_at = time.monotonic()
			crow+=1
			tagscnt-=1
			cdata[t] = {}
			#add new tags
			elq.click()
			if opts.clipboard == '1':
				pyperclip.copy(t)
				time.sleep(PMIN)
				#elq.send_keys(Keys.CONTROL, 'v')
				elq.send_keys(Keys.SHIFT, Keys.INSERT)
			else:
				elq.send_keys(t, Keys.ENTER)
			time.sleep(P100ms*30) # пауза 3

			#get SEO score
			elm=driver.find_element_by_class_name('stat-value-seo-score')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('seo {}'.format(elm1.text))
			cdata[t]['seo'] = float(elm1.text)
			
			if cdata[t]['seo'] < 0.01:
				# если не успело значение обновиться, то подождем еще и вычитаем повторно
				time.sleep(P100ms*20) # пауза 2
				#get SEO score
				elm=driver.find_element_by_class_name('stat-value-seo-score')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('seo2 {}'.format(elm1.text))
				cdata[t]['seo'] = float(elm1.text)
			getScores(driver, cdata[t])
			
			# удалить тег
			try:
				btn = driver.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
				if len(btn) > 1:
					di = driver.find_element_by_id('clear-button')
					di.click()
				else:
					btn[0].click()
			except:
				logging.info("ERROR: {}".format(traceback.format_exc()))
			
			if cdata[t]['ranked'] > 0:
				# если тег ранжированный, то к оценке добавим 0.01
				cdata[t]['seo'] += 0.01
			if t.upper() in vid_title.upper():
				# если тег входит в строку названия, то к оценке добавим 0.02
				cdata[t]['seo'] += 0.02
			
			proctime = time.monotonic() - started_at
			# посчитать среднее время
			times.append(proctime)
			tavg = statistics.median_high(times)
			# sm = 0
			# for tm in times:
			# 	sm += tm
			# tavg = sm / len(times)
			logging.info('видео {:03d}/{:03d}, тег {:03d}/{:03d}, осталось {}, {:03d} {}->{}'.format(indx, urlslen, crow, tagslen, 
				time.strftime('%H:%M:%S', time.gmtime(tavg*tagslen*(urlslen - indx) + (tavg * tagscnt))), tagscnt, t, cdata[t]['seo'] ))	
			#	time.strftime('%H:%M:%S', time.gmtime(proctime*tagslen*(urlslen - indx) + (proctime * tagscnt))), tagscnt, t, cdata[t]['seo'] ))
			
			saveindb({'vid': vid, 'url': cururl, 'tags': {t: cdata[t]}, })
				
		#logging.info('cancel changes')
		btn=driver.find_element_by_id('discard')
		btn.click()
		time.sleep(PMIN)
	return 33

def importTags(opts):
	logging.info(f'Загрузка списка видео из таблицы TagImport')
	(urls, tags_cnt, tag_max_len) = getVidTags4Import(opts.dt)
	urlslen = len(urls.keys())
	logging.info('Подготовлено {} видео для анализа'.format(urlslen))
	tagslen = tags_cnt
	logging.info('Подготовлено {} тегов для анализа'.format(tagslen))
	
	driver = connect2Browser(opts)
	indx = 0
	# проход по видео
	for vid, curvid in urls.items():
		indx+=1
		exitOnKey()
		cururl = curvid["url"]
		#tagsIN = curvid["tags"]
		tags = curvid["tags"]
		logging.info(f'{indx:03d}/{urlslen:03d} открытие страницы {vid} видео {cururl}')
		checkPauseKey() # key p
		cdata = {}
		BIGlen = 0
		rankedCNT = 0
		
		#tags = getNewTags4Vid2(cururl, tagsIN, opts.dt)
		tagslen = len(tags)

		logging.info(f'{tagslen} тегов для оценки под текущим видео с учетом даты {opts.dt}')
		if tagslen < 1:
			logging.info('Для видео все теги содержат актуальные оценки. Переходим к следующему видео.')
			continue # перейти к рассмотрению тегов следующего видео
		driver.get(cururl)
		time.sleep(P100ms*20) # пауза 2
		
		try:
			testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
		except TimeoutException:
			logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
			continue
				
		inpts = driver.find_elements_by_xpath("//ytcp-mention-input")
		vid_title = inpts[0].text
		logging.info(f'Наименование видео {vid_title}')
		# добавить два слова для очистки тегов, чтобы гарантированно отображалась кнопка удалить все теги
		elq = driver.find_element_by_id('text-input')
		elq.click()
		# #clear all tags
		doit_clear=2
		while doit_clear>0:
			try:
				di = driver.find_element_by_id('clear-button')
				di.click()
				doit_clear=0
			except:
				logging.info("find_element_by_id('clear-button') Unexpected error: {}, {}, {}".format(sys.exc_info()[0], sys.exc_info()[1], traceback.format_exc()))
				elq.send_keys('wr1,wr2', Keys.ENTER)
				time.sleep(PMIN)
				doit_clear-=1

		# по тегам сохранить рейтинг
		tagscnt=tagslen
		crow = 0
		times = []
		for tagelm in tags:
			t = tagelm['t']
			ttype = tagelm['ttype']
			checkPauseKey() # key p
			started_at = time.monotonic()
			crow+=1
			tagscnt-=1
			cdata[t] = {}
			#add new tags
			elq.click()
			if opts.clipboard == '1':
				pyperclip.copy(t)
				time.sleep(PMIN)
				#elq.send_keys(Keys.CONTROL, 'v')
				elq.send_keys(Keys.SHIFT, Keys.INSERT)
			else:
				elq.send_keys(t, Keys.ENTER)
			time.sleep(P100ms*30) # пауза 3

			#get SEO score
			elm=driver.find_element_by_class_name('stat-value-seo-score')
			elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
			#logging.info('seo {}'.format(elm1.text))
			cdata[t]['seo'] = float(elm1.text)
			cnt_retry = 1
			while cdata[t]['seo'] < 0.01 and cnt_retry < 5:
				cnt_retry +=1
				# если не успело значение обновиться, то подождем еще и вычитаем повторно
				time.sleep(P100ms*20) # пауза 2
				#get SEO score
				elm=driver.find_element_by_class_name('stat-value-seo-score')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('seo2 {}'.format(elm1.text))
				cdata[t]['seo'] = float(elm1.text)
			getScores(driver, cdata[t])
			
			# удалить тег
			try:
				btn = driver.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
				if len(btn) > 1:
					di = driver.find_element_by_id('clear-button')
					di.click()
				else:
					btn[0].click()
			except:
				logging.info("ERROR: {}".format(traceback.format_exc()))
			
			extstr = '|'
			if cdata[t]['seo'] > 7:
				BIGlen += len(t)
				extstr += '  BIG '
			else:
				extstr += ' '*6
			extstr += '{:04d}'.format(BIGlen)
			if (cdata[t]['ranked'] > 0) and (cdata[t]['seo'] > 11.79):
				# если тег ранжированный, то к оценке добавим 0.01
				cdata[t]['seo'] += 0.01
				rankedCNT += 1
				extstr += ' RANKED '
			else:
				extstr += ' '*8
			extstr += '{}'.format(rankedCNT)
			if (t.upper() in vid_title.upper()) and (cdata[t]['seo'] > 7.0):
				# если тег входит в строку названия, то к оценке добавим 0.02
				cdata[t]['seo'] += 0.02
			
			proctime = time.monotonic() - started_at
			# посчитать среднее время
			times.append(proctime)
			tavg = statistics.median_high(times)
			# sm = 0
			# for tm in times:
			# 	sm += tm
			# tavg = sm / len(times)
			infostr='видео {:03d}/{:03d}, тег {:03d}/{:03d}, осталось {}, {:03d} {}->{:05.2f} '.format(indx, urlslen, crow, tagslen, 
				time.strftime('%H:%M:%S', time.gmtime(tavg*tagslen*(urlslen - indx) + (tavg * tagscnt))), tagscnt, t, cdata[t]['seo'])
			infostr+=' '*(60+tag_max_len - len(infostr))
			logging.info(f'{infostr}{extstr}')	
			#	time.strftime('%H:%M:%S', time.gmtime(proctime*tagslen*(urlslen - indx) + (proctime * tagscnt))), tagscnt, t, cdata[t]['seo'] ))
			
			saveindb({'vid': vid, 'url': cururl, 'tags': {t: cdata[t]}, })
			if opts.slim =='1':
				if (BIGlen >= 510) and (rankedCNT > 4):
					logging.info('Для видео {} длина тегов BIG {}, RANKED {}. Переходим к следующему видео.'.format(cururl, BIGlen, rankedCNT))
					break #gotoNextVideo
				
		#logging.info('cancel changes')
		btn=driver.find_element_by_id('discard')
		btn.click()
		time.sleep(PMIN)
	return 34

def getTitleRest(opts):
	logging.info('Загрузка списка видео из БД таблица CheckData')
	urls = loadUrlsDB()
	urlslen = len(urls.keys())
	logging.info('Подготовлено {} видео для анализа'.format(urlslen))
	indx = 0
	# проход по видео
	logging.info(f'№;Кол-во;ИД видео;Ссылка;Наименование;"Остаток наименования"')
	for vid, curv in urls.items():
		indx+=1
		curtitle = curv['title']
		cururl = curv['url']
		titlerest = removeTags(vid, cururl, curtitle)
		logging.info(f'{indx:03d};{urlslen:03d};{vid};{cururl};"{curtitle}";"{titlerest}"')
	return 13

def loadUrlsDB():
	cdata = {}
	with orm.db_session:
		for cd in CheckData.select(lambda t: t.data['title'] is not None):
			vid = cd.data['url'].split('/')[4]
			cdata[vid] = {'url'  : cd.data['url']
					  ,'title': cd.data['title']}
	return cdata

def removeTags(vid, url, title):
	resstr = title.lower()
	lasttag = ''
	with orm.db_session:
		for t in orm.select(t.tag for t in TagSEO).distinct().order_by(orm.raw_sql('LENGTH(tag) DESC')):
			resstr = resstr.replace(t.lower(), '')
		
		resstr = ''.join([ch for ch in resstr if ch not in '.,;:!?'])
		resstr = ''.join([ch for ch in resstr if not ch.isdigit()])
		ext = ['эффективный'
			  ,'эффективные'
			  ,'эффективная'
			  ,'эффективной'
			  ,'эффективными'
			  ,'эффективному'
			  ,'эффективное'
			  ,' из '
			  ,' и '
			  ]
		for e in ext:
			resstr = resstr.replace(e, ' ')

		todo = 100
		while todo > 1:
			curstr = resstr.replace('  ', ' ')
			if resstr != curstr:
				todo -=1
				resstr = curstr
			else:
				todo = 0
		resstr = resstr.strip()
	return resstr

def PrepareTags(opts):
	# если требуется очистить таблицу
	if opts.truncate == '1':
		with orm.db_session:
			db.execute('DELETE FROM TagImport;')
			#db.execute('VACUUM;')
	# по списку видео из файла или по всем видео в таблице
	if len(opts.infile) > 1:
		# read videos from infile
		urls=loadUrls(opts.infile)
	else:
		urls = {}
		# read videos from table TagSEO
		with orm.db_session:
			for v in orm.select((vtag.vid, vtag.url) for vtag in TagSEO):
				urls[v[0]] = v[1]
	urlslen = len(urls.keys())
	logging.info('Подготовлено {} видео для анализа'.format(urlslen))

	rtags = [] # список ранжированных тегов
	rvtags = {} # список ранжированных тегов с группировкой под видео
	atags = [] # список тегов для проверки в названии
	ztags = [] # список нулевых тегов
	zvtags = {}
	ctags = [] # список рабочих тегов
	cvtags = {}

	# добавочные теги
	if len(opts.addtags) > 2:
		tagsIN=loadTags(opts.addtags)
		tagslen = len(tagsIN)

	# ранжированные теги
	if opts.rtag == '1': # сформировать списки ранжированных с группировкой под видео
		lasturl = '-'
		with orm.db_session:
			vrtags = orm.select((x.vid, x.url, x.tag) for x in TagSEO if x.seo > 11.79 and x.ranked > 0).distinct().order_by(orm.raw_sql("vid"))
			lenrtags = len(vrtags)
			for qt in vrtags:
				cururl = qt[1]
				if lasturl == cururl:
					rvtags[cururl].append(qt[2])
				else:
					rvtags[cururl] = [qt[2]]
					lasturl = cururl
			logging.info(f'Ранжированных тегов под видео {lenrtags}')
			#logging.info(f'{rvtags}')
	elif opts.rtag == '2': # сформировать список ранжированных тегов по всей таблице
		with orm.db_session:
			rtags = orm.select(x.tag for x in TagSEO if x.seo > 11.79 and x.ranked > 0)
			lenrtags = len(rtags)
			logging.info(f'Ранжированных тегов {lenrtags}')

	# все теги для выявления вхождения в название
	if opts.ttag in '12': # не применимо, в любом случае необходимо сформировать список слов из всех тегов
		with orm.db_session:
			atags = orm.select(tt.tag for tt in TagSEO).distinct().order_by(orm.raw_sql("LENGTH(tag) desc"))
			lenrtags = len(atags)
			logging.info(f'Уникальных в текущей таблице тегов {lenrtags}')

	# нулевые теги
	if opts.ztag == '1': # сформировать списки нулевых с группировкой под видео
		lasturl = '-'
		with orm.db_session:
			vztags = orm.select((x.vid, x.url, x.tag) for x in TagSEO if x.seo >= MaxScore or x.seo < MinScore).distinct().order_by(orm.raw_sql("vid"))
			lenztags = len(vztags)
			for qt in vztags:
				cururl = qt[1]
				if lasturl == cururl:
					zvtags[cururl].append(qt[2])
				else:
					zvtags[cururl] = [qt[2]]
					lasturl = cururl
			logging.info(f'Нулевых тегов под видео {lenztags}')
			#logging.info(f'{zvtags}')
	elif opts.ztag == '2': # сформировать список ранжированных тегов по всей таблице
		with orm.db_session:
			ztags = orm.select(x.tag for x in TagSEO if x.seo >= MaxScore or x.seo < MinScore)
			lenztags = len(ztags)
			logging.info(f'Нулевых тегов {lenztags}')
	
	# рабочие теги			
	if opts.ctag == '1': # сформировать списки нулевых с группировкой под видео
		lasturl = '-'
		with orm.db_session:
			vctags = orm.select((x.vid, x.url, x.tag) for x in TagSEO if x.seo > WrkScore and x.seo < MaxScore).distinct().order_by(orm.raw_sql("vid"))
			lenсtags = len(vctags)
			for qt in vctags:
				cururl = qt[1]
				if lasturl == cururl:
					cvtags[cururl].append(qt[2])
				else:
					cvtags[cururl] = [qt[2]]
					lasturl = cururl
			logging.info(f'Рабочих тегов под видео {lenсtags}')
			#logging.info(f'{cvtags}')
	elif opts.ctag == '2': # сформировать список ранжированных тегов по всей таблице
		with orm.db_session:
			ctags = orm.select(x.tag for x in TagSEO if x.seo > WrkScore and x.seo < MaxScore)
			lenсtags = len(ctags)
			logging.info(f'Рабочих тегов {lenсtags}')


	# пройдем по видео
	cnt = 0
	started_at_video = time.monotonic()
	#for v in urls:
	for v, cururl in urls.items():
		cnt+=1
		intags = []
		#cururl = v #v[1]
		
		logging.info('--------------------------------------------------')
		resttime = (urlslen - cnt) * (time.monotonic() - started_at_video) / cnt
		resttimestr = time.strftime('%H:%M:%S', time.gmtime(resttime))
		logging.info(f'{cnt}/{urlslen}, осталось {resttimestr}, Вставка тегов для видео {cururl}')
		with orm.db_session:
			#vidtags = [tg.tag for tg in TagSEO.select(lambda x: x.url==cururl)]
			# прочитать название видео
			cd = CheckData.get(url=cururl)
			if cd is not None:
				vidtitle = cd.data['title'].lower()
			else:
				vidtitle = '-'
				logging.info(' -- Для анализа названия видео требуется выполнить анализ Чек-листа')
			logging.info(f'название видео: {vidtitle}')
			
			if len(opts.addtags) > 2:
				cntinsert = 0
				for t in tagsIN:
					ti = TagImport.get(url=cururl, tag=t)
					if ti is None:
						TagImport(url=cururl, tag=t, ttype='ADDED')
					cntinsert+=1
				logging.info(f'Вставлено добавочных тегов: {cntinsert}')
			
			if opts.rtag in '12':
				cntinsert = 0
				if opts.rtag == '1':
					rtags = rvtags[cururl] if cururl in rvtags.keys() else []
				with orm.db_session:
					for t in rtags:
						# проверить вхождение слов из тега в название видео
						if checkTagWordsInTitle(t,vidtitle, opts.wordsintitle):
							ti = TagImport.get(url=cururl, tag=t)
							if ti is None:
								TagImport(url=cururl, tag=t, ttype='RANKED')
						cntinsert+=1

				logging.info(f'Вставлено ранжированных тегов: {cntinsert}')

			if opts.ztag in '12':
				cntinsert = 0
				if opts.ztag == '1':
					ztags = zvtags[cururl] if cururl in zvtags.keys() else []
				with orm.db_session:
					for t in ztags:
						ti = TagImport.get(url=cururl, tag=t)
						if ti is None:
							TagImport(url=cururl, tag=t, ttype='ZERO')
						cntinsert+=1
				logging.info(f'Вставлено нулевых тегов: {cntinsert}')

			if opts.ctag in '12':
				cntinsert = 0
				if opts.ctag == '1':
					ctags = cvtags[cururl]  if cururl in cvtags.keys() else []
				with orm.db_session:
					for t in ctags:
						ti = TagImport.get(url=cururl, tag=t)
						if ti is None:
							TagImport(url=cururl, tag=t, ttype='CLOUD')
						cntinsert+=1
				logging.info(f'Вставлено рабочих тегов: {cntinsert}')

			if opts.ttag in '12':
				if vidtitle =='-':
					logging.info(f'Требуется запустить режим сбора чек-листа для видео {cururl}')
				else:
					cntinsert = 0
					vidtitle = vidtitle.lower()
					with orm.db_session:
						for tt in atags:
							tg = tt.lower()
							#logging.info(f'{tg} в {vidtitle}')
							if tg in vidtitle:
								ti = TagImport.get(url=cururl, tag=tt)
								if ti is None:
									TagImport(url=cururl, tag=tt, ttype='TITLED')
								vidtitle = vidtitle.replace(tg, '')
								cntinsert += 1
							if len(vidtitle) < 4:
								break
					logging.info(f'Вставлено тегов входящих в название: {cntinsert}')


	return 1

def checkTagWordsInTitle(tag,vidtitle, wit='0'):
	res = False
	if wit == '0':
		res = True
	else:
		# check
		cnt = 0
		vttl = vidtitle.lower()
		# tag split on words
		wrds = tag.lower().split(' ')
		# count words in title
		for w in wrds:
			if w in vttl:
				cnt+=1
		if cnt >= int(wit):
			res = True

	return res 

def saveAnalytPg(opts):
	(wb, wss, wsd, wst, newWBfile) = tags_openxls(opts.infile)
	tags_clearStatSheet(opts, wsd)
	urls = anl_getVUrls(wss, opts.owner)

	urlslen = len(urls)
	logging.info(f'---------- Очередь для обработки {urlslen} ссылок')

	#odt = datetime.datetime.strptime(opts.dt, '%Y-%m-%d %H:%M')
	#logging.info('---- дата свежести сохранненых данных {}'.format(odt))
	if urlslen > 0:
		driver = connect2Browser(opts)
	crow=1
	for u in urls:
		logging.info('-------- Обработка {:03d} из {:03d} ({:05.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen,u['title']))
		crow+=1
		#logging.info(u['url'])
		driver.get(u['url'])
		time.sleep(P100ms*20) # пауза 2
		try:
			testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_xpath('//h1[@class="page-title style-scope ytcp-app"][1]'))
		except TimeoutException:
			logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
			continue
		anl_savePg(driver, 'Обзор', u['videoId'])
		tabs = driver.find_elements_by_xpath('//ytcp-ve[@class="style-scope yta-screen"]')
		if tabs:
			tabs[1].click()
			time.sleep(P100ms*10)
			anl_savePg(driver, 'Просмотры', u['videoId'])
			tabs[2].click()
			time.sleep(P100ms*10)
			anl_savePg(driver, 'Взаимодействие', u['videoId'])
			tabs[3].click()
			time.sleep(P100ms*10)
			anl_savePg(driver, 'Аудитория', u['videoId'])
	return 1

def anl_savePg(drv, pgTab, videoId):
	pth = './pages'
	Path(pth).mkdir(parents=True, exist_ok=True)
	flname = '{}/{}_{}_{}.html'.format(pth, FLTYPE[pgTab]['litera'], videoId, datetime.datetime.now().strftime('%y%m%d_%H%M%S'))
	with open(flname, 'w', encoding='utf-8') as flres:
			flres.writelines([drv.page_source])
	logging.info(f'Сохранена закладка {pgTab} в файле {flname}')
def parseAnalytPg(opts):
	pass

def getSEOtags_V2(vid):
	tgs=[]
	tranked = []
	with orm.db_session:
		for tg in TagSEO.select(lambda t: t.vid==vid and t.real > 7.0).order_by(orm.raw_sql('real desc, length(tag)')):
			#if tg.tag not in tgs:
			tgs+=[tg.tag]
			if tg.ranked > 0:
				tranked.append(tg.tag)
	return (tgs, tranked)

def getRankedTags_V2(vid):
	tgs=[]
	with orm.db_session:
		for tg in TagSEO.select(lambda t: t.vid==vid and t.real > 6.7 and t.real < 7.1 and t.ranked == 1).order_by(orm.raw_sql('real desc, seo desc, length(tag)')):
			#if tg.tag not in tgs:
			tgs+=[tg.tag]
	return tgs

def yt_len(pstr):
	plen = 0
	for ch in pstr:
		plen+=2 if ch==' ' else 1
	return plen

def tagsUpdate_V2(drv, vid, url, o):
	# вторая попытка улучшить алгоритм обновления облака тегов
	svd = 0 # не привело к сохранению в видео ютюб
	drv.get(url)
	time.sleep(P100ms*20) # пауза 2
	try:
		testElm = WebDriverWait(drv, float(o.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
	except TimeoutException:
		logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
		
	# сформировать предлагаемый перечень тегов для замены
	(ntags, nrtags) = getSEOtags_V2(vid) # теги для видео удовлетворяющие условию отбора
	logging.info('Ранжированных {} в общем облаке из {} тегов'.format(len(nrtags), len(ntags)))
	rtags = getRankedTags_V2(vid) # ранжированные теги для видео
	#curtags = []

	# сохранить текущее состояние видео
	vidSEO1 = getYTseo4Vid(drv)
	logging.info('seo1: {}, real: {}, show: {}'.format(vidSEO1['seo'], vidSEO1['treal'], vidSEO1['tshow']))
	
	rtagStr = ''
	rtagslen = 0
	rcnt = 0
	if len(rtags) > 0:
		# начать формирование строки тегов
		for t in rtags:
			tlen = yt_len(t)
			if rtagslen < 500 and rtagslen+tlen < 500 and rcnt < int(o.rtags):
				rcnt += 1
				rtagStr += ','+t
				rtagslen +=  tlen + 1 #запятая
			else:
				break
		logging.info('тегов ранжированных {}, длина {}'.format(rcnt, rtagslen))
	
	tagStr = ''
	tagslen = 0
	tcnt = 0
	if tagslen < 540 and len(ntags) > 0 :
		# дополнить строку тегов
		for t in ntags:
			tlen = yt_len(t)
			if tagslen+tlen+1 < 520 - rtagslen:
				tcnt += 1
				tagStr += ','+t
				tagslen +=  tlen + 1 #запятая
			else:
				break
		logging.info('тегов общих {}, длина {}'.format(tcnt, tagslen))
		tagStr += ',' + rtagStr
		tagslen = yt_len(tagStr)
		logging.info('тегов {}, длина {} '.format(tcnt+rcnt, tagslen))
		#tagStr += ',техно1,техно2' # перед вставкой добавляем для получения новых результатов плагина
		inp = drv.find_element_by_id('text-input')
		clearAlltags(drv, inp)
		inp.click()
		pyperclip.copy(tagStr)
		time.sleep(PMIN)
		# вставить новую строку тегов
		#inp.send_keys(Keys.CONTROL, 'v')
		inp.send_keys(Keys.SHIFT, Keys.INSERT)
		time.sleep(P100ms*20)
		# проверить длинну <div slot="description" id="tags-count" class="style-scope ytcp-video-metadata-basics">155/500</div>
		tlen = 501
		while tlen > 500:
			# откорректировать длинну
			tagslength = drv.find_element_by_id('tags-count')
			#logging.info('tags-count: {}'.format(tagslength.text))
			chcnt = tagslength.text.split('/')
			#logging.info('tags-count: {}'.format(chcnt[0]))
			tlen = int(chcnt[0])
			logging.info(f'длина облака тегов {tlen}')
			if tlen > 500:
				delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
				if len(delbtns) > 0:
					delbtns[-(rcnt+1)].click()
	else:
		logging.info('Нет слов для формирования строки тегов. Работаем с текущими')
		# вставлять нечего, далее работаем с тем, что есть

	# получить текущее состояние видео
	time.sleep(P100ms*30)
	vidSEO2 = getYTseo4Vid(drv)
	logging.info('seo2: {}, real2: {}, show2: {}'.format(vidSEO2['seo'], vidSEO2['treal'], vidSEO2['tshow']))
	
	#если рейтинг выше, то сохранить
	if (vidSEO2['treal'] > 49.99) or ((vidSEO2['treal'] > vidSEO1['treal']) and (vidSEO1['tshow']<0.1 or vidSEO2['tshow'] > vidSEO1['tshow'])):
		logging.info('!!! {:05.2f} >= {:05.2f} SAVE, seo2: {}, seo1: {}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['seo'], vidSEO1['seo'], vidSEO2['tshow'], vidSEO1['tshow']))
		# save this
		svd = 1
		elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
		if len(elms)>0:
			elms[0].click()
			time.sleep(PMIN)
	
	# продолжаем улучшать, пробуем удалять по одному тегу, возможно найдем более высокое значение рейтинга
	todo = True if vidSEO2['treal'] < 50 else False
	fup = 0
	fdown = 0
	iteration = 0
	vidSEOlast = vidSEO2.copy()
	while todo and iteration < 100 and fdown < 1 and fup < 2:
		iteration+=1
		tlen = readTagsLen(drv)
		deltrue = False
		deltag = ''
		if tlen > 0:
			#пытаемся удалить последний тег
			delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
			if len(delbtns) > 0:
				parentelm = delbtns[-1].find_element_by_xpath("./..")
				deltag = parentelm.get_attribute('vidiq-keyword')
				delbtns[-1].click()
				time.sleep(P100ms*20)
				deltrue = True

		vidSEOcur = getYTseo4Vid(drv)
		if deltrue:
			logging.info('удаляем тег: {}, seo: {}, real: {}, show: {}'.format(deltag,
						vidSEOcur['seo'], vidSEOcur['treal'], vidSEOcur['tshow']))

		if ( vidSEOlast['treal'] > vidSEOcur['treal'] ): #and ( abs(vidSEOlast['tshow'] - vidSEOcur['tshow']) <0.1 ):
			# уменьшился рейтинг после удаления тега - отменяем и завершаем подбор
			fdown +=1
			logging.info('--- {:05.2f} < {:05.2f} DISCARD'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal'])))
			vidSEOlast = vidSEOcur.copy()

		elif ( vidSEOlast['treal'] < vidSEOcur['treal']) and (vidSEO1['treal'] <= vidSEOcur['treal']):
			# сохранить и продолжить удалять до следующего изменения
			svd = 1
			fup +=1
			logging.info('!!! {:05.2f} > {:05.2f} SAVE {}'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal']), fup))
			elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
			if len(elms)>0:
				elms[0].click()
				time.sleep(PMIN)
			todo = True if vidSEOcur['treal'] < 50 else False
			vidSEOlast = vidSEOcur.copy()

	if fdown > 0:
		# cancel
		#logging.info(' :( discard, seo2 {:05.2f} =< {:05.2f} seo1.'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
		btn=drv.find_element_by_id('discard')
		btn.click()
		

	# сохранить рейтинг в БД
	time.sleep(P100ms*30) #пауза 3
	vidSEO2 = getYTseo4Vid(drv)
	if float(vidSEO2['treal']) > float(vidSEO1['treal']):
		logging.info('!!! текущий seo: {:05.2f} > старого {:05.2f}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['tshow'], vidSEO1['tshow']))
	elif float(vidSEO2['treal']) <= float(vidSEO1['treal']):
		logging.info('текущий seo: {:05.2f} <= старому {:05.2f}'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
	
	saveSEOupdate(vid, [vidSEO1, vidSEO2], svd)

def tagsUpdate_V3(drv, vid, url, o):
	# добавляем в конец тэг с оценкой выше 7.0 для планового удаления и обновления пересчета SEO
	# вторая попытка улучшить алгоритм обновления облака тегов
	svd = 0 # не привело к сохранению в видео ютюб
	drv.get(url)
	time.sleep(P100ms*20) # пауза 2
	try:
		testElm = WebDriverWait(drv, float(o.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
	except TimeoutException:
		logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
		
	# сформировать предлагаемый перечень тегов для замены
	(ntags, nrtags) = getSEOtags_V2(vid) # теги для видео удовлетворяющие условию отбора
	logging.info('Ранжированных {} в общем облаке из {} тегов'.format(len(nrtags), len(ntags)))
	rtags = getRankedTags_V2(vid) # ранжированные теги для видео
	#curtags = []

	# сохранить текущее состояние видео
	vidSEO1 = getYTseo4Vid(drv)
	logging.info('seo1: {}, real: {}, show: {}'.format(vidSEO1['seo'], vidSEO1['treal'], vidSEO1['tshow']))
	
	rtagStr = ''
	rtagslen = 0
	rcnt = 0
	if len(rtags) > 0:
		# начать формирование строки тегов
		for t in rtags:
			tlen = yt_len(t)
			if rtagslen < 500 and rtagslen+tlen < 500 and rcnt < int(o.rtags):
				rcnt += 1
				rtagStr += ','+t
				rtagslen +=  tlen + 1 #запятая
			else:
				break
		logging.info('тегов ранжированных {}, длина {}'.format(rcnt, rtagslen))
	
	tagStr = ''
	tagslen = 0
	tcnt = 0
	if tagslen < 540 and len(ntags) > 0 :
		# дополнить строку тегов
		for t in ntags:
			tlen = yt_len(t)
			if tagslen+tlen+1 < 510 - rtagslen:
				tcnt += 1
				tagStr += ','+t
				tagslen +=  tlen + 1 #запятая
			else:
				rtagStr += ','+t
				rcnt += 1
				break
		logging.info('тегов общих {}, длина {}'.format(tcnt, tagslen))
		tagStr += ',' + rtagStr
		tagslen = yt_len(tagStr)
		logging.info('тегов {}, длина {} '.format(tcnt+rcnt, tagslen))
		#tagStr += ',техно1,техно2' # перед вставкой добавляем для получения новых результатов плагина
		inp = drv.find_element_by_id('text-input')
		clearAlltags(drv, inp)
		inp.click()
		pyperclip.copy(tagStr)
		time.sleep(PMIN)
		# вставить новую строку тегов
		#inp.send_keys(Keys.CONTROL, 'v')
		inp.send_keys(Keys.SHIFT, Keys.INSERT)
		time.sleep(P100ms*20)
		# проверить длинну <div slot="description" id="tags-count" class="style-scope ytcp-video-metadata-basics">155/500</div>
		tlen = 501
		while tlen > 500:
			# откорректировать длинну
			tagslength = drv.find_element_by_id('tags-count')
			#logging.info('tags-count: {}'.format(tagslength.text))
			chcnt = tagslength.text.split('/')
			#logging.info('tags-count: {}'.format(chcnt[0]))
			tlen = int(chcnt[0])
			logging.info(f'длина облака тегов {tlen}')
			if tlen > 500:
				delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
				if len(delbtns) > 0:
					delbtns[-(rcnt+1)].click()
	else:
		logging.info('Нет слов для формирования строки тегов. Работаем с текущими')
		# вставлять нечего, далее работаем с тем, что есть

	# получить текущее состояние видео
	time.sleep(P100ms*30)
	cnt_retry = 1
	vidSEO2 = getYTseo4Vid(drv)
	while vidSEO2['seo'] < 0.01 and cnt_retry < 5:
		cnt_retry += 1
		# если не успело значение обновиться, то вставляем тег, ожидаем, удаляем, читаем повторно
		inp = drv.find_element_by_id('text-input')
		inp.click()
		pyperclip.copy('экономика')
		time.sleep(PMIN)
		inp.send_keys(Keys.SHIFT, Keys.INSERT)
		time.sleep(P100ms * 30)  # пауза 2
		delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
		if len(delbtns) > 0:
			delbtns[-1].click()
		time.sleep(P100ms * 30)  # пауза 2
		# get SEO score
		vidSEO2 = getYTseo4Vid(drv)

	#vidSEO2 = getYTseo4Vid(drv)
	logging.info('seo2: {}, real2: {}, show2: {}'.format(vidSEO2['seo'], vidSEO2['treal'], vidSEO2['tshow']))
	
	#если рейтинг выше, то сохранить
	if (vidSEO2['treal'] > 49.99) or ((vidSEO2['treal'] > vidSEO1['treal']) and (vidSEO1['tshow']<0.1 or vidSEO2['tshow'] > vidSEO1['tshow'])):
		logging.info('!!! {:05.2f} >= {:05.2f} SAVE, seo2: {}, seo1: {}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['seo'], vidSEO1['seo'], vidSEO2['tshow'], vidSEO1['tshow']))
		# save this
		svd = 1
		elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
		if len(elms)>0:
			elms[0].click()
			time.sleep(PMIN)
	
	# продолжаем улучшать, пробуем удалять по одному тегу, возможно найдем более высокое значение рейтинга
	todo = True if vidSEO2['treal'] < 50 else False
	fup = 0
	fdown = 0
	iteration = 0
	vidSEOlast = vidSEO2.copy()
	while todo and iteration < 100 and fdown < 1 and fup < 2:
		iteration+=1
		tlen = readTagsLen(drv)
		deltrue = False
		deltag = ''
		if tlen > 0:
			#пытаемся удалить последний тег
			delbtns = drv.find_elements_by_xpath("//ytcp-icon-button[@id='delete-icon']")
			if len(delbtns) > 0:
				parentelm = delbtns[-1].find_element_by_xpath("./..")
				deltag = parentelm.get_attribute('vidiq-keyword')
				delbtns[-1].click()
				time.sleep(P100ms*20)
				deltrue = True

		vidSEOcur = getYTseo4Vid(drv)
		if deltrue:
			logging.info('удаляем тег: {}, seo: {}, real: {}, show: {}'.format(deltag,
						vidSEOcur['seo'], vidSEOcur['treal'], vidSEOcur['tshow']))

		if ( vidSEOlast['treal'] > vidSEOcur['treal'] ): #and ( abs(vidSEOlast['tshow'] - vidSEOcur['tshow']) <0.1 ):
			# уменьшился рейтинг после удаления тега - отменяем и завершаем подбор
			fdown +=1
			logging.info('--- {:05.2f} < {:05.2f} DISCARD'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal'])))
			vidSEOlast = vidSEOcur.copy()

		elif ( vidSEOlast['treal'] < vidSEOcur['treal']) and (vidSEO1['treal'] <= vidSEOcur['treal']):
			# сохранить и продолжить удалять до следующего изменения
			svd = 1
			fup +=1
			logging.info('!!! {:05.2f} > {:05.2f} SAVE {}'.format(float(vidSEOcur['treal']), float(vidSEOlast['treal']), fup))
			elms=drv.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
			if len(elms)>0:
				elms[0].click()
				time.sleep(PMIN)
			todo = True if vidSEOcur['treal'] < 50 else False
			vidSEOlast = vidSEOcur.copy()

	if fdown > 0:
		# cancel
		#logging.info(' :( discard, seo2 {:05.2f} =< {:05.2f} seo1.'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
		btn=drv.find_element_by_id('discard')
		btn.click()
		

	# сохранить рейтинг в БД
	time.sleep(P100ms*30) #пауза 3
	vidSEO2 = getYTseo4Vid(drv)
	if float(vidSEO2['treal']) > float(vidSEO1['treal']):
		logging.info('!!! текущий seo: {:05.2f} > старого {:05.2f}, show2: {}, show1: {}'.format(float(vidSEO2['treal']), float(vidSEO1['treal']), vidSEO2['tshow'], vidSEO1['tshow']))
	elif float(vidSEO2['treal']) <= float(vidSEO1['treal']):
		logging.info('текущий seo: {:05.2f} <= старому {:05.2f}'.format(float(vidSEO2['treal']), float(vidSEO1['treal'])))
	
	saveSEOupdate(vid, [vidSEO1, vidSEO2], svd)

def BackupZeroTags(opts):
	logging.info('Подготавливаем перемещение нулевых тегов в архивную таблицу')
	cnt = 0
	with orm.db_session():
		ztags = TagSEO.select(lambda x: x.seo < WrkScore).order_by(TagSEO.url)
		ztagslen = len(ztags)
		logging.info(f'Отобрано для перемещения в архивную таблицу {ztagslen}')
		for zt in ztags:
			cnt +=1
			#logging.info(f'vid={zt.vid}, tag={zt.tag}')
			indx=0
			for at in TagSEOArch.select(lambda x: x.vid==zt.vid and x.tag==zt.tag):
				indx+=1
				at.delete()
				#logging.info(f'{indx} vid={at.vid}, tag={at.tag}')
			at = TagSEOArch(dt=zt.dt, vid=zt.vid, url=zt.url,
							tag=zt.tag, seo=zt.seo, real=zt.real,
							tcount=zt.tcount, tpopular=zt.tpopular, tintitle=zt.tintitle,
							tindesc=zt.tindesc, triple=zt.triple, tshow=zt.tshow,
							ranked=zt.ranked, hivolume=zt.hivolume, data=zt.data
							)
			# else:
			# 	if zt.dt > at.dt:
			# 		at.set(dt = zt.dt,
			# 				seo = zt.seo,
			# 				real = zt.real,
			# 				tcount = zt.tcount,
			# 				tpopular = zt.tpopular,
			# 				tintitle = zt.tintitle,
			# 				tindesc = zt.tindesc,
			# 				triple = zt.triple,
			# 				tshow = zt.tshow,
			# 				ranked = zt.ranked,
			# 				hivolume = zt.hivolume,
			# 				data = zt.data
			# 			)
			zt.delete()
			if str(cnt)[-1] == '0':
				logging.info(f'удаленных записей {cnt}')
				orm.flush()
	logging.info(f'удаленных записей {cnt}')
	return 1
def RestoreZeroTags(opts):
	logging.info('Подготавливаем перемещение нулевых тегов в основную таблицу')
	cnt = 0
	with orm.db_session():
		ztags = TagSEOArch.select().order_by(TagSEOArch.url)
		ztagslen = len(ztags)
		logging.info(f'Отобрано для перемещения в основную таблицу {ztagslen}')
		for zt in ztags:
			cnt +=1
			#logging.info(f'vid={zt.vid}, tag={zt.tag}')
			indx=len(TagSEO.select(lambda x: x.vid==zt.vid and x.tag==zt.tag))
			if indx == 0:
				at = TagSEO(dt=zt.dt, vid=zt.vid, url=zt.url,
								tag=zt.tag, seo=zt.seo, real=zt.real,
								tcount=zt.tcount, tpopular=zt.tpopular, tintitle=zt.tintitle,
								tindesc=zt.tindesc, triple=zt.triple, tshow=zt.tshow,
								ranked=zt.ranked, hivolume=zt.hivolume, data=zt.data
								)
			# else:
			# 	if zt.dt > at.dt:
			# 		at.set(dt = zt.dt,
			# 				seo = zt.seo,
			# 				real = zt.real,
			# 				tcount = zt.tcount,
			# 				tpopular = zt.tpopular,
			# 				tintitle = zt.tintitle,
			# 				tindesc = zt.tindesc,
			# 				triple = zt.triple,
			# 				tshow = zt.tshow,
			# 				ranked = zt.ranked,
			# 				hivolume = zt.hivolume,
			# 				data = zt.data
			# 			)
			zt.delete()
			if str(cnt)[-1] == '0':
				logging.info(f'восстановленных записей {cnt}')
				orm.flush()
	logging.info(f'Завершено. Восстановленных записей {cnt}')
	return 1

def testfunc(opts):
	
	k = 0
	buf = ''

	while k < 10:
		k+=1
		print(f'next step {k}')
		time.sleep(P100ms*30)
		try:
			if msvcrt.kbhit():
				key = msvcrt.getch()
				print(key)   # just to show the result
				if key ==b'p':
					s = input('Pause. Press any key and ENTER to continue...')
		except:
			print(traceback.format_exc())
	return 

def checkPauseKey(k=b'p'):
	try:
		if msvcrt.kbhit():
			key = msvcrt.getch()
			#print(key)   # just to show the result
			if key == k:
				s = input('Pause. Press any key and ENTER to continue...')
	except:
		print(traceback.format_exc())

def exitOnKey(k=b'q'):
	try:
		if msvcrt.kbhit():
			key = msvcrt.getch()
			#print(key)   # just to show the result
			if key == k:
				logging.info('Нажата секретная клавиша. Принудительный выход из программы. {}', format(datetime.datetime.now()))
				exit(99)
	except:
		print(traceback.format_exc())

def getVideoIDfromURL(url):
	#TEST: выделить ИД видео из ссылки на видео
	(videoID, videoURL) = getVID(url)
	return videoID

# простановка лайков по списку видео для активного пользователя, только для тех видео которые отсутствуют в таблице данных Like2Video
def like2video(opts):
	urls = []
	#DONE: прочитать список видео из opts.infile
	urls = loadVideoUrlsFromTxt(opts.infile)
	urlslen = len(urls)
	#DONE: подключиться к браузеру
	driver = connect2Browser(opts)

	#DONE: получить текущего пользователя ютюб
	driver.get('https://youtube.com')
	time.sleep(P100ms*20) # пауза 2 при заходе на очередное видео
	try:
		testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_xpath('//button[@class="style-scope ytd-topbar-menu-button-renderer"][1]'))
	except TimeoutException:
		logging.info("Превышено время ожидания загрузки страницы. Попытка обработать следующую ссылку.")
				
	avatar_btn= driver.find_elements_by_xpath('//button[@class="style-scope ytd-topbar-menu-button-renderer"][1]')
	avatar_btn[0].click()
	time.sleep(PMIN)
	avatar_name = driver.find_elements_by_id('account-name')[0].text
	#DONE: идти по списку видео, проверяя отсутствия записи в таблице
	crow=0
	for u in urls:
		crow+=1
		exitOnKey() # key q
		logging.info('-------- Обработка {:03d} из {:03d} ({:06.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen, u["url"]))
		checkPauseKey() # key p
		
		vidID = getVideoIDfromURL(u["url"])
		likeslen = 0
		with orm.db_session():
			#DONE: прочитать из БД наличие лайка
			likes = Like2Video.select(lambda x: x.vid==vidID and x.auser==avatar_name)
			likeslen = len(likes)
		
			if likeslen == 0:
				#DONE: загрузить страницу видео
				driver.get(u["url"])
				time.sleep(P100ms*20) # пауза 2 при заходе на очередное видео
				#DONE: прочитать текущее состояние кнопки Лайк, если уже установлен лайк, записать в БД
				like_btn= driver.find_elements_by_xpath('//a[@class="yt-simple-endpoint style-scope ytd-toggle-button-renderer"]')[0]
				#$x('//a[@class="yt-simple-endpoint style-scope ytd-toggle-button-renderer"]')[2].click()
				#DONE: если лайк можно поставить, то поставить
				#$x('//a[@class="yt-simple-endpoint style-scope ytd-toggle-button-renderer"][1]//button[@id="button" and @aria-pressed="false"]')[0]
				btns = like_btn.find_elements_by_xpath('.//button[@id="button"]')
				if btns[0].get_attribute('aria-pressed') == 'false':
					#like_btn.click()
					driver.find_elements_by_xpath('//a[@class="yt-simple-endpoint style-scope ytd-toggle-button-renderer"]')[2].click()
				#DONE: зафиксировать в БД установку лайка
				newLike = Like2Video(vid=vidID, auser=avatar_name)
				orm.flush()
				logging.info('-------- поставили Лайк')
			else:
				logging.info('-------- Лайк уже установлен')
		#DONE: перейти к следующему видео в списке или завершить работу
	return 1

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('--timeout', help='waiting timeout to load the page', default='10')
	parser.add_argument('--infile', help='input file xlsm', default='_')
	parser.add_argument('--webdriver', help='web driver path', default='C:\\bin\\Selenium\\chromedriver')
	parser.add_argument('--update', help='update new and zero tags', default='0')
	# 0- сбор тегов для входных видео; 1-новые и с оценкой ноль теги добавить для проверки; 2-проверка тегов из других видео на seo под анализируемым видео
	parser.add_argument('--clipboard', help='inserts tags by clipboard', default='1')
	
	parser.add_argument('--dt', help='expire datetime', default='2020-05-12 12:00')
	parser.add_argument('--words', help='rate words on seo in infile', default='0')
	
	parser.add_argument('--chID', help='channel ID', default='-')
	parser.add_argument('--chvideos', help='only channels videos', default='0')
	parser.add_argument('--outflname', help='part of the filename', default='chvideos')
	parser.add_argument('--apiKey', help='google api key', default='-')
	parser.add_argument('--tags', help='expire datetime', default='-')
	parser.add_argument('--seo',help='seo to process', default='50')
	parser.add_argument('--owner',help='owner of videos', default='Олег Брагинский')
	parser.add_argument('--analyt',help='analytics page scenario', default='-')
	parser.add_argument('--add',help='adds tags and estimate on videos', default='-')
	parser.add_argument('--addtags',help='file of tags', default='-')
	parser.add_argument('--test',help='test func', default='-')
	parser.add_argument('--truncate',help='truncate TagImport', default='-')
	parser.add_argument('--rtag',help='ranked tags for TagImport', default='-')
	parser.add_argument('--ztag',help='zero and unnotmaly hi tags for TagImport', default='-')
	parser.add_argument('--ctag',help='cloud tags for TagImport', default='-')
	parser.add_argument('--ttag',help='tags in video title for TagImport', default='-')
	parser.add_argument('--rtags',help='ranked tags count', default='-')
	parser.add_argument('--arch',help='process operations backup/restore zero tags', default='-')
	parser.add_argument('--slim',help='break tag estimation when big >510 and count of ranked == 5', default='0')
	parser.add_argument('--wordsintitle',help='ranked tag will be estimated if wordsintitle words exists in title string', default='0')
	parser.add_argument('--xls',help='to process excel file', default='-')
	parser.add_argument('--short',help='to process excel file', default='100')
	parser.add_argument('--PMS',help='base pause value', default='0.20')
	parser.add_argument('--likes', help='process likes', default='-')

	args = parser.parse_args()
	started_at = time.monotonic()
	print(args.arch)
	if args.PMS:
		P100ms = float(args.PMS)
	print(f'P100ms={P100ms}')
	if args.likes == '1':
		like2video(args)
	elif args.tags == '0':
		check_list(args)
	elif args.tags == '1':
		set_tags(args)
	elif args.tags == '2':
		getTitleRest(args)
	elif args.analyt == '1':
		saveAnalytPg(args)
	elif args.analyt == '2':
		parseAnalytPg(args)
	elif args.add == '1':
		addAndEstimate(args)
	elif args.add == '2':
		importTags(args) # go 1378
	elif args.add == '3':
		PrepareTags(args) # go 1628
	elif args.arch == 'b':
		BackupZeroTags(args) # go 1990
	elif args.arch == 'r':
		RestoreZeroTags(args) # go 
	elif args.test == '1':
		testfunc(args)
	else:
		main(args)
	proctime = time.monotonic() - started_at
	logging.info(f'Выполнение сценария завершено за {proctime:.3f} сек.')
	# cmd line python yt_optima.py --timeout=12 --webdriver="C:\Windows\chromedriver.exe" --infile="!in.txt"