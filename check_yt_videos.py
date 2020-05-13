# -*- coding: utf-8 -*-
import sys
import argparse
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import * #

import pyperclip
import os
import time
import datetime
from check_yt_models import *
from pony import orm

import logging
logging.basicConfig()
logging.getLogger().setLevel(logging.INFO)
logging.getLogger().addHandler(logging.FileHandler("info_yt_videos.log"))


def main(opts):
	logging.info(opts.xls)
	#logging.info(os.path.dirname(opts.xls))
	wb = openpyxl.load_workbook(filename=opts.xls, read_only=False, keep_vba=True)
	if 'Обновление' in wb.sheetnames:
		wss = wb['Обновление']
	else:
		logging.info('source worksheet did not find')
		exit(2)
	if 'Статистика' in wb.sheetnames:
		wsd = wb['Статистика']
	else:
		logging.info('destination worksheet did not find')
		exit(3)
	if 'Исключения' in wb.sheetnames:
		wse = wb['Исключения']
	else:
		logging.info('exclusions worksheet did not find')
	if 'Теги' in wb.sheetnames:
		wst= wb['Теги']
	else:
		logging.info('tags worksheet did not find')

	newWBfile = '{}\\Ссылки_{}.xlsm'.format(os.path.dirname(opts.xls), datetime.datetime.now().strftime('%y%m%d_%H%M'))
	logging.info(newWBfile)
	
	# Chrome driver
	chrome_options = Options()

	chrome_options.add_argument('--start-maximized') # чтобы поместилась панель vidIQ
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222") # запустить предварительно Хром
	driver = webdriver.Chrome(opts.chrome, options=chrome_options)

	time.sleep(3)
	driver.set_page_load_timeout(120)
	driver.implicitly_wait(3)
	driver.maximize_window()
	#протестировать подключение к хрому
	driver.get('https://youtube.com')
	try:
		testElm = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_id("avatar-btn"))
	except TimeoutException:
		logging.info("Выполните вход в аккаунт ютюб и перезапустите приложение.")
		exit(13)

	if opts.tags == '0': # обновление статистики чек-листа
		urls = getUrls(wss, wse)
		urlslen = len(urls)
		logging.info('----------urls len {}, {}'.format(urlslen, repr(urls[0])))

		odt = datetime.datetime.strptime(opts.dt, '%Y-%m-%d %H:%M')
		logging.info('---- дата свежести сохранненых данных {}'.format(odt))
		crow=1
		for u in urls:
			logging.info('-------- Обработка {:03d} из {:03d} ({:05.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen,u['title']))
			crow+=1

			svdata = CheckSavedData(u['url'], odt)
			if svdata is not None and float(opts.seo) <= float(svdata["seo"]):
				# записать в эксель данные из БД
				write2xls(wsd, crow, svdata)
			else:
				#запустить сбор данных и сохранение для дальнейшего использования
				if svdata is not None:
					logging.info('в БД seo = {}'.format(svdata["seo"]))
				else:
					logging.info('новое видео в списке, сбор информации')
				driver.get(u['url'])
				wsd.cell(row=crow, column=1, value=u['num'])
				wsd.cell(row=crow, column=2, value=u['title'])
				wsd.cell(row=crow, column=3, value=u['url'])
				time.sleep(1)
				# Обрабатывать ошибку "чужое видео" пример https://studio.youtube.com/video/ez5hxF1jP-A/edit
				try:
					myElem = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
				except TimeoutException:
					if driver.find_element_by_id("error-image"):
						logging.info("Видео не доступно. {}, url={}".format(u['title'], u['url']))
						continue
					logging.info("Loading took too much time!")

				driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				
				cdata={'num': u['num'],
					   'title': u['title'],
					   'url': u['url'],}

				#get SEO score
				elm=driver.find_element_by_class_name('stat-value-seo-score')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				logging.info('seo {}'.format(elm1.text))
				wsd.cell(row=crow, column=4, value=float(elm1.text))
				cdata['seo'] = elm1.text
				
				#get h1 stat-value-undefined two times
				elms=driver.find_elements_by_class_name('stat-value-undefined')
				#logging.info('elms count {}'.format(len(elms)))
				elm1 = elms[0].find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('undefined {}'.format(elm1.text))
				wsd.cell(row=crow, column=5, value=float(elm1.text))
				cdata['stat-value1'] = elm1.text

				elm1 = elms[1].find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('undefined {}'.format(elm1.text))
				wsd.cell(row=crow, column=11, value=int(elm1.text))
				cdata['stat-value2'] = elm1.text

				#get stat-value-tag-count
				elm=driver.find_element_by_class_name('stat-value-tag-count')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('count {}'.format(elm1.text))
				wsd.cell(row=crow, column=6, value=int(elm1.text))
				cdata['tag-count'] = elm1.text

				#get stat-value-tag-volume
				elm=driver.find_element_by_class_name('stat-value-tag-volume')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('volume {}'.format(elm1.text))
				wsd.cell(row=crow, column=7, value=int(elm1.text))
				cdata['tag-volume'] = elm1.text

				#get stat-value-keywords-in-title
				elm=driver.find_element_by_class_name('stat-value-keywords-in-title')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				#logging.info('title {}'.format(elm1.text))
				wsd.cell(row=crow, column=8, value=int(elm1.text))
				cdata['keywords-in-title'] = elm1.text

				#get stat-value-keywords-in-description
				try:
					elm=driver.find_element_by_class_name('stat-value-keywords-in-description')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					#logging.info('seo {}'.format(elm1.text))
					wsd.cell(row=crow, column=9, value=int(elm1.text))
					cdata['keywords-in-description'] = elm1.text
				except:
					logging.info("stat-value-keywords-in-description Unexpected error: {}".format(sys.exc_info()[0]))

				#get stat-value-tripled-keywords
				try:
					elm=driver.find_element_by_class_name('stat-value-tripled-keywords')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					#logging.info('tripled-keywords {}'.format(elm1.text))
					wsd.cell(row=crow, column=10, value=int(elm1.text))
					cdata['tripled-keywords'] = elm1.text
				except:
					logging.info("stat-value-tripled-keywords Unexpected error: {}".format(sys.exc_info()[0]))

				
				#get <span class="stat-value stat-value-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
				try:
					elm=driver.find_element_by_class_name('stat-value-ranked-tags')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					#logging.info('undefined {}'.format(elm1.text))
					wsd.cell(row=crow, column=12, value=int(elm1.text))
					cdata['ranked-tags'] = elm1.text
				except:
					logging.info("stat-value-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

				#get <span class="stat-value stat-value-high-volume-ranked-tags"><span class="value-inner">0</span><span class="out-of">/5</span></span>
				try:
					elm=driver.find_element_by_class_name('stat-value-high-volume-ranked-tags')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					#logging.info('undefined {}'.format(elm1.text))
					wsd.cell(row=crow, column=13, value=int(elm1.text))
					cdata['volume-ranked-tags'] = elm1.text
				except:
					logging.info("stat-value-high-volume-ranked-tags Unexpected error: {}".format(sys.exc_info()[0]))

				#get checklist
				try:
					driver.find_element_by_class_name('stat-box-checklist').click()
					elm=driver.find_element_by_class_name('stat-value-checklist')
					#logging.info(elm.get_attribute('innerHTML'))
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					logging.info('check-list {}'.format(elm1.text))
					col=14
					wsd.cell(row=crow, column=col, value=int(elm1.text))
					cdata['chl-14'] = int(elm1.text)

					elm = driver.find_element_by_class_name('seo-checklist')
					for el in elm.find_elements_by_xpath(".//li")[:9]:
						col+=1
						wsd.cell(row=crow, column=col, value=1 if 'checked' == el.get_attribute('class') else 0)
						cdata['chl-{}'.format(col)] = 1 if 'checked' == el.get_attribute('class') else 0
				except:
					logging.info("stat-box-checklist Unexpected error: {}".format(sys.exc_info()[0]))

				SaveCheckData(cdata)
			# except:
			# 	continue
		wb.save(newWBfile)
	elif opts.tags == '1': #подставляем теги
		print('------------------tags')
		urltags = getUrlsTags(wss, wse, wst)
		print('------------------urls count ', len(urltags))
		#logging.info('key:{}, title:{}, tags:{}'.format(urltags[0]['key'], urltags[0]['title'], urltags[0]['tags']))
		crow=1
		urlslen = len(urltags)
		for u in urltags:
			logging.info('-------- Обработка {:03d} из {:03d} ({:06.2f} %), {}'.format(crow, urlslen, 100.00*crow/urlslen,u['title']))
			crow+=1

			svdata = None #CheckSavedData(u['url'], odt)
			if svdata is not None:
				# записать в эксель данные из БД
				#write2xls(wsd, crow, svdata)
				pass
			else:
				#запустить установку тегов
				driver.get(u['url'])
				time.sleep(1)
				# Обрабатывать ошибку "чужое видео" пример https://studio.youtube.com/video/ez5hxF1jP-A/edit
				try:
					myElem = WebDriverWait(driver, float(opts.timeout)).until(lambda x: x.find_element_by_class_name("stat-value-high-volume-ranked-tags"))
				except TimeoutException:
					if driver.find_element_by_id("error-image"):
						logging.info("Видео не доступно. {}, url={}".format(u['title'], u['url']))
						continue
					logging.info("Loading took too much time!")

				driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				
				cdata={'num': u['num'],
					   'title': u['title'],
					   'url': u['url'],}

				#get SEO1 score
				elm=driver.find_element_by_class_name('stat-value-seo-score')
				elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
				logging.info('seo1 {}'.format(elm1.text))
				cdata['seo1'] = float(elm1.text)
				
				# Обновлять только если текущее значение seo меньше заданного граничного из эксель
				#logging.info('seo {} vs rate {}'.format(cdata['seo1'], float(u['taginfo']['maxseo'])))
				if cdata['seo1'] < float(u['taginfo']['maxseo']):

					elq = driver.find_element_by_id('text-input')
					elq.click()
					pyperclip.copy('word1, word2')
					time.sleep(0.5)
					elq.send_keys(Keys.SHIFT, Keys.INSERT)
					time.sleep(0.5)
					#clear all tags
					try:
						di = driver.find_element_by_id('clear-button')
						di.click()
					except:
						logging.info("find_element_by_id('clear-button') Unexpected error: {}".format(sys.exc_info()[0]))

					#add new tags
					elq.click()
					pyperclip.copy(u['taginfo']['newtags'])
					time.sleep(0.5)
					elq.send_keys(Keys.SHIFT, Keys.INSERT)
					time.sleep(3)

					#get SEO score
					elm=driver.find_element_by_class_name('stat-value-seo-score')
					elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
					logging.info('seo2 {}'.format(elm1.text))
					cdata['seo2'] = float(elm1.text)
					
					if cdata['seo2'] < 0.01:
						# если не успело значение обновиться, то подождем еще и вычитаем повторно
						time.sleep(2)
						#get SEO score
						elm=driver.find_element_by_class_name('stat-value-seo-score')
						elm1 = elm.find_elements_by_xpath(".//span[@class='value-inner']")[0]
						logging.info('seo2 {}'.format(elm1.text))
						cdata['seo2'] = float(elm1.text)

					if cdata['seo2'] > cdata['seo1']:
						# примем решение сохранить если значение увеличилось
						logging.info('save')
						elms=driver.find_elements_by_xpath("//ytcp-button[@id='save']") #driver.find_element_by_id('save')
						if len(elms)>0:
							elms[0].click()
							time.sleep(1)
					else:
						logging.info('cancel')
						elm=driver.find_element_by_id('discard')
						elm.click()
						time.sleep(0.5)
				else:
					logging.info('seo {} > tag rate {}'.format(cdata['seo1'], float(u['taginfo']['maxseo'])))	
				
				#exit(998) #удалить перед релизом, только для теста

def getUrls(ws, wse):
	exurls = []
	urls=[]
	doit = True
	r = 1
	emptyrows = 0
	# сформировать список исключений
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = wse.cell(row=r, column=6).value
		if curval is not None:
			if 'youtube.com' in curval:
				exurls+=[curval]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	# прочитать ссылки на ютюб для сбора информации по чег-лист
	doit = True
	r = 1
	emptyrows = 0
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = ws.cell(row=r, column=6).value
		if curval is not None:
			if ('youtube.com' in curval) and (not curval in exurls):
				url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/edit'
				logging.info(url)
				urls+=[{'url': url, 'num': ws.cell(row=r, column=1).value, 'title': ws.cell(row=r, column=5).value}]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
		

	return urls

def getUrlsTags(ws, wse, wst):
	titlemask = {} # маски названий
	exurls = [] # исключения
	urls=[] # видео

	doit = True
	r = 1
	emptyrows = 0
	# сформировать список исключений
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = wse.cell(row=r, column=6).value
		if curval is not None:
			if 'youtube.com' in curval:
				exurls+=[curval]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	
	
	doit = True
	r = 1
	emptyrows = 0
	# сформировать список масок названий
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = wst.cell(row=r, column=2).value
		if curval is not None:
			if '%' in curval:
				curkey = curval.replace('%', '')
				titlemask[curkey]={'newtags': wst.cell(row=r, column=4).value, 'maxseo': wst.cell(row=r, column=3).value, }
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False
	#logging.info(titlemask.keys())
	# прочитать ссылки на ютюб для установки тегов
	doit = True
	r = 1
	emptyrows = 0
	while doit:# and r < 3: #для анализа всего экселя требуется оставить только doit
		r+=1
		curval = ws.cell(row=r, column=6).value
		title = ws.cell(row=r, column=5).value
		if curval is not None and title is not None:
			if ('youtube.com' in curval) and (not curval in exurls):
				for k in titlemask.keys():
					if k in title:
						url = curval.replace('https://youtube.com/watch?v=', 'https://studio.youtube.com/video/')+'/edit'
						logging.info(url)
						urls+=[{'url': url, 'num': ws.cell(row=r, column=1).value, 'title': ws.cell(row=r, column=5).value, 'key': k, 'taginfo': titlemask[k], }]
		else:
			emptyrows+=1
		if emptyrows > 3:
			doit = False

	return urls

def CheckSavedData(url, dt):
	cdjson = None
	with orm.db_session:
		cd = CheckData.get(url=url)
		if cd is not None:
			if cd.dt > dt:
				cdjson = cd.data
	return cdjson

def SaveCheckData(cdata):
	with orm.db_session:
		cd = CheckData.get(url=cdata['url'])
		if cd is None:
			cd = CheckData(url=cdata['url'], data=cdata)
		else:
			cd.data=cdata
			cd.dt=datetime.datetime.now()

def write2xls(ws, crow, data):
	ws.cell(row=crow, column=1, value=int(data["num"]) if 'num' in data else '')
	ws.cell(row=crow, column=2, value=data["title"] if 'title' in data else '')
	ws.cell(row=crow, column=3, value=data["url"] if 'url' in data else '')

	ws.cell(row=crow, column=4, value=float(data["seo"]) if 'seo' in data else '')

	ws.cell(row=crow, column=5, value=float(data["stat-value1"]) if 'stat-value1' in data else '')
	ws.cell(row=crow, column=6, value=int(data["tag-count"]) if 'tag-count' in data else '')
	ws.cell(row=crow, column=7, value=int(data["tag-volume"]) if 'tag-volume' in data else '')
	ws.cell(row=crow, column=8, value=int(data["keywords-in-title"]) if 'keywords-in-title' in data else '')
	ws.cell(row=crow, column=9, value=int(data["keywords-in-description"]) if 'keywords-in-description' in data else '')
	ws.cell(row=crow, column=10, value=int(data["tripled-keywords"]) if 'tripled-keywords' in data else '')
	ws.cell(row=crow, column=11, value=int(data["stat-value2"]) if 'stat-value2' in data else '')
	ws.cell(row=crow, column=12, value=int(data["ranked-tags"]) if 'ranked-tags' in data else '')
	ws.cell(row=crow, column=13, value=int(data['volume-ranked-tags']) if 'volume-ranked-tags' in data else '')
	
	for i in range(14,24):
		key = 'chl-{}'.format(i)
		ws.cell(row=crow, column=i, value=int(data[key]) if key in data else '')


if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	#parser.add_argument('--dt', help='report date "YYYY-MM-DD"', default=datetime.datetime.today().strftime('%Y-%m-%d'))
	parser.add_argument('--timeout', help='waiting timeout to load the page', default='10')
	parser.add_argument('--xls', help='input file xlsm', default='c:\\web\\GAPIpy3\\yt_data_v3\\Ссылки_200501_2300.xlsm')
	parser.add_argument('--chrome', help='chrome driver path', default='C:\\Web\\GAPIpy3\\yt_data_v3\\driver\\chromedriver')
	parser.add_argument('--chromedir', help='chrome user profile data path', default='C:\\UpDate\YT\\test_data')
	parser.add_argument('--dt', help='expire datetime', default='2020-05-03 12:00')
	parser.add_argument('--tags', help='expire datetime', default='0', const='0', nargs='?')
	parser.add_argument('--seo',help='seo to process', default='50')
	args = parser.parse_args()
	main(args)
	